# ============================================================
#  VRAMaぴょん 🐾 by くろぴな
#  ・誰がVRAMを食ってるか「見る」     （プロセス別VRAM一覧）
#  ・そこだけピンポイントで「手放す」 （Ollamaモデル解放／プロセス終了）
#
#  ポイント:
#   nvidia-smi は GeForce + Windows だとプロセス別VRAMが [N/A] になる。
#   なので Windows の性能カウンター(GPU Process Memory)から
#   PIDごとの専用VRAM使用量を取る。これが今回のキモ。
# ============================================================

import os
import re
import sys
import json
import shutil
import winreg
import subprocess
import threading
import urllib.request
from collections import deque

# ── 外部依存はGUIのcustomtkinterだけ（Ollama通信はstdlibのurllibで完結）──
try:
    import customtkinter as ctk
except ImportError:
    print("【準備】customtkinter をインストールするね...")
    subprocess.run([sys.executable, "-m", "pip", "install", "customtkinter"], check=True)
    import customtkinter as ctk

from tkinter import messagebox, Menu, Canvas

# ─────────────────────────────────────────────────────────
#  設定
# ─────────────────────────────────────────────────────────
OLLAMA_URL   = "http://localhost:11434"   # Ollamaのアドレス（必要なら変えてね）
FORGE_URL    = "http://127.0.0.1:7860"     # Stable Diffusion WebUI Forge（--api起動が必要）
COMFY_URL    = "http://127.0.0.1:8188"     # ComfyUI（未導入でもOK＝繋がらなければ無効表示）
REFRESH_MS   = 2500                        # 自動更新の間隔(ミリ秒)
NOWIN        = 0x08000000 if os.name == "nt" else 0  # コンソール窓を出さない

# 🔍「詳しく診断」で起動する別アプリ（Ollama診断GUI）の場所さがし。
#   フォルダ配置はマシンごとに違う（同梱/隣/工房…）ので、相対パス1本に頼らず
#   候補を順に探す。どこにも無ければ🔍はグレースフルに「見つからないよ」表示。
DIAG_FILENAME = "VRAMaPyonDiagnostics.py"
# 旧名（後方互換の候補。診断ツールのフォルダ/ファイル旧称でも拾えるように）
DIAG_DIRNAMES = ["VRAMaぴょん診断(VRAMaPyonDiagnostics)",
                 "Ollama ボトルネック診断ツール v2 by くろぴな"]
DIAG_FILENAMES = [DIAG_FILENAME, "ollama_diagnosis_gui.py"]

def find_diagnosis_tool():
    """VRAMaぴょん診断の.pyを候補順に探して、最初に見つかった絶対パスを返す（無ければ""）。

    フォルダの★リネームや改名に強いよう、最後にファイル名のグロブ探索でフォールバックする。
    """
    import glob
    here = os.path.dirname(os.path.abspath(__file__))
    # ① 環境変数で明示指定（フルパス or フォルダ）＝確実な上書き手段
    env = os.environ.get("VRAMAPYON_DIAG", "").strip().strip('"')
    if env:
        p = os.path.normpath(env if env.lower().endswith(".py")
                             else os.path.join(env, DIAG_FILENAME))
        if os.path.exists(p):
            return p
    # ② 既知の場所×既知の名前（同梱 here/・here/diag・隣・2つ上）
    bases = [here, os.path.join(here, "diag"),
             os.path.join(here, ".."), os.path.join(here, "..", "..")]
    for base in bases:
        for fn in DIAG_FILENAMES:
            p = os.path.normpath(os.path.join(base, fn))             # 同梱フラット
            if os.path.exists(p):
                return p
        for d in DIAG_DIRNAMES:
            for fn in DIAG_FILENAMES:
                p = os.path.normpath(os.path.join(base, d, fn))      # フォルダ入り
                if os.path.exists(p):
                    return p
    # ③ グロブ：兄弟/2つ上の任意フォルダ内に診断.pyがあれば拾う（★や改名に強い）
    for parent in (os.path.join(here, ".."), os.path.join(here, "..", "..")):
        for fn in DIAG_FILENAMES:
            hits = glob.glob(os.path.normpath(os.path.join(parent, "*", fn)))
            if hits:
                return hits[0]
    return ""

# ポップ＆キュート カラートークン（基準: ぷろんぷたんUI）
PINK       = "#EA4C9D"
PINK_DARK  = "#D43A87"
PINK_SOFT  = "#FBE3F1"
LAVENDER   = "#8B7EE8"   # ブランド基準ラベンダー（旧#B79CEDは淡すぎ→白文字が浮いていた）
BG         = "#FFF4FA"
CARD       = "#FFFFFF"
INK        = "#4A3340"
SUB        = "#9B8090"
# しきい値カラー（仕様書の色変化）
C_OK   = "#5FBF7A"   # ～79%  通常
C_WARN = "#F2C744"   # 80%～  注意
C_HIGH = "#F2843C"   # 90%～  警告
C_CRIT = "#E8453C"   # 95%～  危険

# プロセス行: VRAM推移グラフを「文字の後ろ」にうっすら敷く（あしあとライン＋先端に🐾肉球）。
# 左セルを1枚のCanvasにして 背景=グラフ→前面=名前/PID文字 を重ね描き。等間隔で右から流すので
# 行幅に追従＆幅が広くてもスカスカにならない。グラフは淡色＝文字が主役、肉球だけ濃ピンク。
SPARK_STEP    = 5         # 1サンプルあたりの横px（固定間隔）
SPARK_N       = 160       # 履歴サンプル数（直近 N×REFRESH_MS）
CELL_H        = 44        # 左セル(=行の文字＋背景グラフ)の高さ(px)
SPARK_BG_LINE = "#F19BC0" # 背景グラフの線（文字が読めるよう淡め）
SPARK_BG_FILL = "#FBE2EE" # 背景グラフの面（さらに淡く）
SPARK_RED_LINE = "#EE8B86" # 張り付き(危険)時の線
SPARK_RED_FILL = "#FBD9D6" # 張り付き(危険)時の面
SPARK_CEIL    = "#E7B7C7" # GPU比モードの「天井(満タン)」を示す点線
# 足型の状態判定（瞬間のスパイクで🐾がチカチカしないよう“持続”だけ拾う）
PIN_SHARE  = 0.70  # GPU総量のこの割合以上を占有＝大物
PIN_FLAT   = 0.05  # 直近の変動がこの相対幅未満＝高止まり＝天井張り付き
RISE_RATIO = 1.06  # 後半平均が前半平均のこの倍以上＝持続的に増加
RESV_GAP_BYTES = int(0.3 * 1024**3)  # 予約(Dedicated)が実数(Local)よりこれ以上多い行だけ「・予約○G」を併記

APP_FONT  = ("Yu Gothic UI", 13)
APP_BOLD  = ("Yu Gothic UI", 13, "bold")
BIG_FONT  = ("Yu Gothic UI", 22, "bold")
SM_FONT   = ("Yu Gothic UI", 11)

# ─────────────────────────────────────────────────────────
#  言語 / Language（既定=日本語、🌐ボタンで日英トグル）
# ─────────────────────────────────────────────────────────
LANG = "ja"

TR = {
    "ja": {
        "app_title": "VRAMaぴょん 🐾 by くろぴな",
        "app_name": "VRAMaぴょん",
        "header_title": "VRAMaぴょん  🐾",
        "header_sub": "誰がVRAM食ってる？ → そこだけ手放す",
        "lang_btn": "🌐 EN",
        "btn_refresh": "🔄 更新",
        "sw_auto": "自動更新",
        "sw_ai_only": "AIだけ表示",
        "btn_mini": "🐾 小窓",
        "btn_snap": "📸 記録",
        # ゲージ
        "gauge_loading": "GPU情報を取得中…",
        "gauge_na": "⚠ GPU情報が取れないよ（nvidia-smi無し？）",
        "gauge_empty": "-- / -- GB",
        "mark_ok": "🟢 余裕", "mark_warn": "🟡 注意", "mark_high": "🟠 警告", "mark_crit": "🔴 危険",
        "sub_dedshared": "専用 {du:.1f}/{dt:.1f} GB  ・  共有 {su:.1f}/{st:.1f} GB",
        "igpu_loading": "🍃 内蔵 GPU",
        "igpu_suffix": "（内蔵）",
        "igpu_fallback": "内蔵GPU",
        # Ollama
        "oll_title": "🦙 Ollama のモデルを優しく手放す",
        "oll_checking": "確認中…",
        "oll_disconnected": "🔌 未接続（起動してない？）",
        "oll_none": "✅ ロード中モデルなし",
        "oll_loaded": "💭 {n}個ロード中",
        "oll_btn_all": "全部解放",
        "oll_btn_diag": "🔍 詳しく診断",
        "fit_allgpu": "✅ 全部GPU",
        "fit_ramover": "⚠️ RAM溢れ{p:.0f}%",
        "fit_mostram": "🔴 ほぼRAM",
        "btn_release": "解放",
        # SD
        "sd_title": "🎨 画像生成(SD)のVRAMを解放",
        "sd_disconnected": "🎨 画像生成(SD) — 未接続（Forge/ComfyUIを起動すると出るよ）",
        "sd_ckpt": "ckpt: {name}",
        "sd_vramfree": "VRAM空き {gb:.1f}GB",
        "sd_connected": "繋がってるよ",
        "model_unknown": "(モデル不明)",
        # プロセス一覧
        "proc_title": "VRAMを使ってるプロセス（多い順）",
        "proc_note": "※5070＋内蔵の合算",
        "seg_ratio": "📊 全体比率",
        "seg_detail": "📈 細かい上下",
        "lbl_protected": "🔒 保護",
        "btn_kill": "終了",
        "btn_gpu": "🎛GPU",
        "pref_auto": "自動",
        "pref_fixed": "固定→{label}",
        "pid_text": "PID {pid} ・ {gtag}",
        "pid_text_plain": "PID {pid}",
        "ph_noproc": "プロセス別VRAMが取れなかったよ。\n管理者権限で実行すると拾えることがあるよ。",
        "ph_noai": "（AI関連プロセスは見つからなかったよ）",
        "tag_pinned": "🔥張り付き",
        "tag_rising": "▲増加",
        "resv_tag": " ・予約 {gb:.1f}G",
        "proc_unknown": "(不明なプロセス)",
        # 小窓
        "mini_na": "🐾 GPU情報なし",
        "mini_open": "開く",
        "mini_quit": "終了",
        # GPU指定ダイアログ
        "gpu_dlg_title": "使うGPUを選ぶ",
        "gpu_dlg_head": "🎛 {name}",
        "gpu_dlg_current": "いまの選択 → {label}",
        "gpu_opt_5070": "🚀 RTX5070（高パフォーマンス）",
        "gpu_opt_igpu": "🍃 内蔵GPU（省電力・VRAM節約）",
        "gpu_opt_auto": "🪟 自動（Windows任せ）",
        "gpu_dlg_note": ("※この設定はPC再起動でも残るよ（Windowsのグラフィック設定と同じ）。\n"
                         "※反映には対象アプリの再起動が必要。\n"
                         "※「自動」はWindowsまかせ＝GPUが切り替わることあり。\n"
                         "※LLM/SDのCUDA分やdwm等には効かないよ。"),
        "pref_label_auto": "自動", "pref_label_igpu": "内蔵", "pref_label_5070": "5070",
        # 記録
        "snap_dlg_title": "Excel記録に保存（同じファイルを選べば追記されるよ）",
        "snap_default_name": "VRAMaぴょん_記録_{stamp}.xlsx",
        "snap_ft_excel": "Excel ブック", "snap_ft_csv": "CSV",
        "snap_sheet": "VRAM記録",
        "chart_sheet": "推移グラフ",
        "chart_gpu_title": "GPU全体VRAMの推移(GB)",
        "chart_pct_title": "GPU使用率の推移(%)",
        "chart_proc_title": "プロセス別VRAMの推移(GB)",
        "chart_col_time": "時刻", "chart_col_used": "GPU使用GB",
        "chart_col_total": "GPU合計GB", "chart_col_pct": "使用率%",
        "autorec_start": "▶️ 自動記録 開始！",
        "autorec_recording": "⏹ 記録終了！({n})",
        "autorec_dlg_title": "自動記録の保存先を選んでね（一定間隔で追記するよ）",
        "autorec_err": "自動記録でエラーが出たから止めたよ💦\n（記録ファイルをExcelで開いてたら閉じてね）\n{err}",
        "rec_unit": "秒",
        "snap_saving": "保存中…",
        "snap_done": "記録したよ！🎉\n今回 {n} プロセスを追記。\nこのファイルは通算 {total} 行になったよ。\n\n{path}",
        "snap_done_open": "記録したよ！🎉\n今回 {n} プロセスを追記（通算 {total} 行）。\n\n{path}\n\nこのファイルを今すぐ開く？",
        "snap_fail": "保存に失敗しちゃった💦\n{err}{hint}",
        "snap_perm_hint": "\n（このファイルをExcelで開いてたら、閉じてからもう一度試してね）",
        "flt_ai": "AIだけ", "flt_all": "すべて",
        "totals_na": "(取得できず)",
        "win_none": "ウィンドウなし",
        # メッセージ
        "msg_unload_fail": "「{name}」の解放に失敗したよ💦",
        "msg_forge_confirm": "Forgeのチェックポイントを解放する？\n（次に生成すると自動で読み直すよ）",
        "msg_forge_fail": "Forgeの解放に失敗したよ💦（起動してる？ --api要）",
        "msg_comfy_fail": "ComfyUIの解放に失敗したよ💦（起動してる？）",
        "msg_kill_confirm": "「{name}」(PID {pid}) を終了する？\n保存してないデータは消えちゃうかも。",
        "msg_kill_fail": "終了できなかったよ💦（権限不足かも）",
        "msg_diag_notfound": "診断ツールが見つからないよ💦\n{path}",
        "msg_diag_launchfail": "診断ツールを起動できなかったよ💦\n{e}",
        "msg_pref_set": "「{name}」を「{label}」に設定したよ。\nアプリを再起動すると反映されるよ🐾",
        "msg_pref_fail": "設定の書き込みに失敗したよ💦",
        # 種別ラベル
        "kind_ai": "AI", "kind_maybe": "たぶんAI", "kind_other": "その他", "kind_unknown": "不明",
        # 記録の見出し
        "snap_headers": ["取得時刻", "GPU名", "GPU合計VRAM_GB", "GPU使用VRAM_GB", "GPU空きVRAM_GB",
                         "GPU使用率%", "GPU利用率%", "フィルタ", "プロセス名", "種別", "PID",
                         "プロセスVRAM_GB", "GPU乗り", "モニター", "起動時刻", "実行ファイル"],
        "monitor_q": "モニター?", "monitor_n": "モニター{n}", "monitor_primary": "(主)",
    },
    "en": {
        "app_title": "VRAMaPyon 🐾 by Kuropina",
        "app_name": "VRAMaPyon",
        "header_title": "VRAMaPyon  🐾",
        "header_sub": "Who's eating your VRAM? → free just that",
        "lang_btn": "🌐 日本語",
        "btn_refresh": "🔄 Refresh",
        "sw_auto": "Auto",
        "sw_ai_only": "AI only",
        "btn_mini": "🐾 Mini",
        "btn_snap": "📸 Log",
        "gauge_loading": "Reading GPU info…",
        "gauge_na": "⚠ Can't read GPU info (no nvidia-smi?)",
        "gauge_empty": "-- / -- GB",
        "mark_ok": "🟢 Plenty", "mark_warn": "🟡 Caution", "mark_high": "🟠 Warning", "mark_crit": "🔴 Critical",
        "sub_dedshared": "Dedicated {du:.1f}/{dt:.1f} GB  ・  Shared {su:.1f}/{st:.1f} GB",
        "igpu_loading": "🍃 iGPU",
        "igpu_suffix": " (iGPU)",
        "igpu_fallback": "iGPU",
        "oll_title": "🦙 Gently release Ollama models",
        "oll_checking": "Checking…",
        "oll_disconnected": "🔌 Not connected (not running?)",
        "oll_none": "✅ No models loaded",
        "oll_loaded": "💭 {n} loaded",
        "oll_btn_all": "Release all",
        "oll_btn_diag": "🔍 Diagnose",
        "fit_allgpu": "✅ All on GPU",
        "fit_ramover": "⚠️ {p:.0f}% in RAM",
        "fit_mostram": "🔴 Mostly RAM",
        "btn_release": "Release",
        "sd_title": "🎨 Release image-gen (SD) VRAM",
        "sd_disconnected": "🎨 Image-gen (SD) — not connected (start Forge/ComfyUI to show)",
        "sd_ckpt": "ckpt: {name}",
        "sd_vramfree": "VRAM free {gb:.1f}GB",
        "sd_connected": "connected",
        "model_unknown": "(unknown model)",
        "proc_title": "Processes using VRAM (most first)",
        "proc_note": "* 5070 + iGPU combined",
        "seg_ratio": "📊 GPU share",
        "seg_detail": "📈 Fine detail",
        "lbl_protected": "🔒 Protected",
        "btn_kill": "End",
        "btn_gpu": "🎛GPU",
        "pref_auto": "Auto",
        "pref_fixed": "fixed→{label}",
        "pid_text": "PID {pid} ・ {gtag}",
        "pid_text_plain": "PID {pid}",
        "ph_noproc": "Couldn't read per-process VRAM.\nRunning as administrator may help.",
        "ph_noai": "(No AI-related processes found)",
        "tag_pinned": "🔥Pinned",
        "tag_rising": "▲Rising",
        "resv_tag": " ・reserved {gb:.1f}G",
        "proc_unknown": "(unknown process)",
        "mini_na": "🐾 No GPU info",
        "mini_open": "Open",
        "mini_quit": "Quit",
        "gpu_dlg_title": "Choose GPU",
        "gpu_dlg_head": "🎛 {name}",
        "gpu_dlg_current": "Current → {label}",
        "gpu_opt_5070": "🚀 RTX5070 (high performance)",
        "gpu_opt_igpu": "🍃 iGPU (power saving / VRAM saving)",
        "gpu_opt_auto": "🪟 Auto (let Windows decide)",
        "gpu_dlg_note": ("* This setting persists across reboots (same as Windows Graphics settings).\n"
                         "* The target app must be restarted to apply.\n"
                         "* \"Auto\" lets Windows decide; the GPU may switch.\n"
                         "* Doesn't affect CUDA (LLM/SD) memory or dwm etc."),
        "pref_label_auto": "Auto", "pref_label_igpu": "iGPU", "pref_label_5070": "5070",
        "snap_dlg_title": "Save to Excel log (pick the same file to append)",
        "snap_default_name": "VRAMaPyon_log_{stamp}.xlsx",
        "snap_ft_excel": "Excel workbook", "snap_ft_csv": "CSV",
        "snap_sheet": "VRAM Log",
        "chart_sheet": "Trends",
        "chart_gpu_title": "Total GPU VRAM over time (GB)",
        "chart_pct_title": "GPU usage over time (%)",
        "chart_proc_title": "Per-process VRAM over time (GB)",
        "chart_col_time": "Time", "chart_col_used": "GPU used GB",
        "chart_col_total": "GPU total GB", "chart_col_pct": "Usage %",
        "autorec_start": "▶️ Start auto-log",
        "autorec_recording": "⏹ Stop! ({n})",
        "autorec_dlg_title": "Choose where to auto-log (appends at each interval)",
        "autorec_err": "Auto-log error, stopped 💦\n(If the log file is open in Excel, close it.)\n{err}",
        "rec_unit": "s",
        "snap_saving": "Saving…",
        "snap_done": "Logged! 🎉\nAppended {n} processes this time.\nThis file now has {total} rows total.\n\n{path}",
        "snap_done_open": "Logged! 🎉\nAppended {n} processes ({total} rows total).\n\n{path}\n\nOpen this file now?",
        "snap_fail": "Save failed 💦\n{err}{hint}",
        "snap_perm_hint": "\n(If this file is open in Excel, close it and try again.)",
        "flt_ai": "AI only", "flt_all": "All",
        "totals_na": "(unavailable)",
        "win_none": "no window",
        "msg_unload_fail": "Failed to release \"{name}\" 💦",
        "msg_forge_confirm": "Release Forge's checkpoint?\n(It reloads automatically on the next generation.)",
        "msg_forge_fail": "Failed to release Forge 💦 (running? --api required)",
        "msg_comfy_fail": "Failed to release ComfyUI 💦 (running?)",
        "msg_kill_confirm": "End \"{name}\" (PID {pid})?\nUnsaved data may be lost.",
        "msg_kill_fail": "Couldn't end it 💦 (insufficient privileges?)",
        "msg_diag_notfound": "Diagnosis tool not found 💦\n{path}",
        "msg_diag_launchfail": "Couldn't launch the diagnosis tool 💦\n{e}",
        "msg_pref_set": "Set \"{name}\" to \"{label}\".\nRestart the app to apply 🐾",
        "msg_pref_fail": "Failed to write the setting 💦",
        "kind_ai": "AI", "kind_maybe": "maybe AI", "kind_other": "other", "kind_unknown": "unknown",
        "snap_headers": ["Timestamp", "GPU name", "GPU total VRAM_GB", "GPU used VRAM_GB", "GPU free VRAM_GB",
                         "GPU mem use%", "GPU util%", "Filter", "Process", "Kind", "PID",
                         "Process VRAM_GB", "On GPU", "Monitor", "Start time", "Executable"],
        "monitor_q": "Monitor?", "monitor_n": "Monitor{n}", "monitor_primary": "(main)",
    },
}


def t(key, **kw):
    s = TR.get(LANG, TR["ja"]).get(key) or TR["ja"].get(key, key)
    return s.format(**kw) if kw else s

# AI関連プロセスの判定キーワード（小文字で部分一致）
AI_KEYWORDS = [
    "ollama", "llama-server", "llama_server", "llamacpp", "llama-cpp", "llama-box",
    "koboldcpp", "kobold", "comfyui", "comfy", "forge", "sd.next", "sdnext",
    "invokeai", "invoke", "webui", "stable-diffusion", "exllama", "vllm",
    "textgen", "oobabooga", "lmstudio", "lm-studio", "automatic1111",
]
# 「python」系は中身がForge/Comfyのことが多いので“AIかも”扱い
PY_NAMES = {"python", "python3", "pythonw", "py"}

# 終了させちゃダメなWindows重要プロセス（仕様書: 保護機能）
PROTECTED = {
    "system", "registry", "idle", "smss", "csrss", "wininit", "winlogon",
    "services", "lsass", "lsaiso", "svchost", "dwm", "fontdrvhost",
    "explorer", "sihost", "ctfmon", "runtimebroker", "memory compression",
    "secure system", "audiodg", "conhost", "wudfhost",
}

# ─────────────────────────────────────────────────────────
#  データ取得：プロセス別VRAM（Windows性能カウンター）
# ─────────────────────────────────────────────────────────
# Get-Counter で GPU Process Memory の専用使用量を拾い、PID→使用量に集約。
# インスタンス名の luid_..._phys_N から「どのGPUか」も取り、PIDごとにアダプタ別へ分ける。
# さらに Get-Process で名前とパスを付けてJSONで返す。
PS_VRAM = r"""
try { [Console]::OutputEncoding = [System.Text.UTF8Encoding]::new($false) } catch {}
$ErrorActionPreference = 'SilentlyContinue'
# 専用(Dedicated=予約枠)と 実数(Local=実際にカードに乗ってる分)を両方 "pid|luid" 単位で集める。
# ・Dedicated は over-commit で物理超えもありうる「予約」値（NVIDIA Overlay等が巨大に見える正体）
# ・Local は「そのGPUの専用メモリに実際に乗ってる分」＝実数。表示はこっちを主役にする。
$ded = @{}; $loc = @{}
foreach ($s in (Get-Counter '\GPU Process Memory(*)\Dedicated Usage').CounterSamples) {
  if ($s.CookedValue -gt 0 -and
      $s.InstanceName -match 'pid_(\d+)_luid_(0x[0-9a-fA-F]+_0x[0-9a-fA-F]+)_phys') {
    $key = "$($matches[1])|$($matches[2])"
    if ($ded.ContainsKey($key)) { $ded[$key] += [int64]$s.CookedValue } else { $ded[$key] = [int64]$s.CookedValue }
  }
}
foreach ($s in (Get-Counter '\GPU Process Memory(*)\Local Usage').CounterSamples) {
  if ($s.CookedValue -gt 0 -and
      $s.InstanceName -match 'pid_(\d+)_luid_(0x[0-9a-fA-F]+_0x[0-9a-fA-F]+)_phys') {
    $key = "$($matches[1])|$($matches[2])"
    if ($loc.ContainsKey($key)) { $loc[$key] += [int64]$s.CookedValue } else { $loc[$key] = [int64]$s.CookedValue }
  }
}
$pidDed = @{}; $pidLoc = @{}; $pidAdp = @{}
foreach ($k in (@($ded.Keys) + @($loc.Keys) | Select-Object -Unique)) {
  $parts = $k.Split('|'); $procId = [int]$parts[0]; $luid = $parts[1]
  $bd = 0; if ($ded.ContainsKey($k)) { $bd = $ded[$k] }
  $bl = 0; if ($loc.ContainsKey($k)) { $bl = $loc[$k] }
  if ($pidDed.ContainsKey($procId)) { $pidDed[$procId] += $bd } else { $pidDed[$procId] = $bd }
  if ($pidLoc.ContainsKey($procId)) { $pidLoc[$procId] += $bl } else { $pidLoc[$procId] = $bl }
  if (-not $pidAdp.ContainsKey($procId)) { $pidAdp[$procId] = @{} }
  $pidAdp[$procId][$luid] = @{ ded = [int64]$bd; loc = [int64]$bl }
}
$procs = foreach ($procId in $pidDed.Keys) {
  $p = Get-Process -Id $procId -ErrorAction SilentlyContinue
  $nm = $null; $st = $null; $pa = $null
  if ($p) {
    $nm = $p.ProcessName
    try { $st = $p.StartTime.ToString('o') } catch {}
    try { $pa = $p.Path } catch {}
  }
  $adps = foreach ($lu in $pidAdp[$procId].Keys) {
    [pscustomobject]@{ luid = $lu; bytes = $pidAdp[$procId][$lu].ded; lbytes = $pidAdp[$procId][$lu].loc }
  }
  [pscustomobject]@{ pid = $procId; bytes = $pidDed[$procId]; lbytes = $pidLoc[$procId]; name = $nm;
                     start = $st; path = $pa; adapters = @($adps) }
}
# GPUごとの専用/共有(内蔵ゲージ用)も同じPowerShellプロセス内で取得＝spawnを1回にまとめる
$aded = @{}; $ashr = @{}
foreach ($s in (Get-Counter '\GPU Adapter Memory(*)\Dedicated Usage').CounterSamples) {
  if ($s.InstanceName -match 'luid_(0x[0-9a-fA-F]+_0x[0-9a-fA-F]+)_phys') {
    $k = $matches[1]; if (-not $aded.ContainsKey($k)) { $aded[$k] = [int64]0 }; $aded[$k] += [int64]$s.CookedValue
  }
}
foreach ($s in (Get-Counter '\GPU Adapter Memory(*)\Shared Usage').CounterSamples) {
  if ($s.InstanceName -match 'luid_(0x[0-9a-fA-F]+_0x[0-9a-fA-F]+)_phys') {
    $k = $matches[1]; if (-not $ashr.ContainsKey($k)) { $ashr[$k] = [int64]0 }; $ashr[$k] += [int64]$s.CookedValue
  }
}
$akeys = @($aded.Keys) + @($ashr.Keys) | Select-Object -Unique
$adapters = foreach ($k in $akeys) {
  $d = 0; if ($aded.ContainsKey($k)) { $d = $aded[$k] }
  $h = 0; if ($ashr.ContainsKey($k)) { $h = $ashr[$k] }
  [pscustomobject]@{ luid = $k; dedicated = [int64]$d; shared = [int64]$h }
}
[pscustomobject]@{ procs = @($procs); adapters = @($adapters) } | ConvertTo-Json -Depth 4 -Compress
"""


def get_vram():
    """1回のPowerShell呼び出しで「プロセス別VRAM」と「アダプタ別VRAM」を両方取得＝spawn削減。
    戻り値 (procs, adapters)。procs は VRAM 降順。"""
    out = _run_ps(PS_VRAM)
    procs, adapters = [], []
    if not out.strip():
        return procs, adapters
    try:
        data = json.loads(out)
    except Exception:
        return procs, adapters
    procs = data.get("procs") or []
    if isinstance(procs, dict):
        procs = [procs]
    for d in procs:                      # adapters は単一だと dict で来るので配列へ正規化
        a = d.get("adapters")
        if isinstance(a, dict):
            d["adapters"] = [a]
        elif not a:
            d["adapters"] = []
        # 実数(Local)を主役 bytes にすり替え、予約(Dedicated)は dedBytes へ退避。
        # こうすると既存の表示・バー・並び順・足型・記録がそのまま「実数」基準になる
        # （※adapters の bytes は Dedicated のまま残す＝_resolve_gpu_labels のGPU識別に使うため）。
        d["dedBytes"] = d.get("bytes", 0) or 0
        d["bytes"] = d.get("lbytes", 0) or 0
    procs = [d for d in procs if (d.get("bytes", 0) or 0) > 0 or (d.get("dedBytes", 0) or 0) > 0]
    procs.sort(key=lambda d: d["bytes"], reverse=True)
    adapters = data.get("adapters") or []
    if isinstance(adapters, dict):
        adapters = [adapters]
    return procs, adapters


def _find_powershell() -> str | None:
    """PowerShell本体を探す。PATHに無いことがあるので絶対パスも当たる。"""
    p = shutil.which("powershell")
    if p:
        return p
    sysroot = os.environ.get("SystemRoot", r"C:\Windows")
    cand = os.path.join(sysroot, "System32", "WindowsPowerShell", "v1.0", "powershell.exe")
    if os.path.exists(cand):
        return cand
    return shutil.which("pwsh")


_PS_EXE = _find_powershell()


def _run_ps(script: str, timeout=15) -> str:
    if not _PS_EXE:
        return ""
    try:
        r = subprocess.run(
            [_PS_EXE, "-NoProfile", "-NonInteractive", "-Command", script],
            capture_output=True, text=True, encoding="utf-8", errors="replace",
            timeout=timeout, creationflags=NOWIN)
        return r.stdout or ""
    except Exception:
        return ""


def get_proc_vram() -> list:
    """[{pid, bytes, name, start, path, adapters:[{luid,bytes}]}, ...] を返す。VRAM使用量(降順)。"""
    return get_vram()[0]


# ─────────────────────────────────────────────────────────
#  データ取得：GPU全体（nvidia-smi、合計値はこれで十分）
# ─────────────────────────────────────────────────────────
def get_gpu_totals() -> dict | None:
    """name, total/used/free(MB), util(%) を返す。取れなければNone。"""
    try:
        r = subprocess.run(
            ["nvidia-smi",
             "--query-gpu=name,memory.total,memory.used,memory.free,utilization.gpu",
             "--format=csv,noheader,nounits"],
            capture_output=True, text=True, encoding="utf-8", errors="replace",
            timeout=5, creationflags=NOWIN)
        if r.returncode != 0:
            return None
        p = [x.strip() for x in r.stdout.strip().split(",")]
        return dict(name=p[0], total=int(p[1]), used=int(p[2]),
                    free=int(p[3]), util=int(p[4]))
    except Exception:
        return None


# ─────────────────────────────────────────────────────────
#  GPUごとの「全体メモリ使用率」（Task Managerの“メモリ”相当）
#   nvidia-smiはNVIDIAしか映らない → 内蔵GPU(AMD/Intel)は
#   Windows性能カウンター \GPU Adapter Memory(*)\Dedicated/Shared Usage で拾う。
#   分母(総容量)＝専用VRAM総量(レジストリ qwMemorySize) + 共有上限(物理RAMの半分)。
# ─────────────────────────────────────────────────────────
PS_GPU_STATIC = r"""
$ErrorActionPreference = 'SilentlyContinue'
$ram = (Get-CimInstance Win32_ComputerSystem).TotalPhysicalMemory
$cls = 'HKLM:\SYSTEM\CurrentControlSet\Control\Class\{4d36e968-e325-11ce-bfc1-08002be10318}'
$ad = foreach ($it in (Get-ChildItem $cls)) {
  $p = Get-ItemProperty $it.PSPath
  if ($p.DriverDesc -and $p.'HardwareInformation.qwMemorySize') {
    $mem = $(try { [int64]$p.'HardwareInformation.qwMemorySize' } catch { [int64]0 })
    [pscustomobject]@{ desc = $p.DriverDesc; dedtotal = $mem }
  }
}
[pscustomobject]@{ ram = [int64]$ram; adapters = @($ad) } | ConvertTo-Json -Depth 3 -Compress
"""

_gpu_static = None


def get_gpu_static() -> dict:
    """物理RAM総量と、各表示アダプタの専用VRAM総容量(レジストリ)。内蔵GPUゲージの分母用。一度だけ取得しキャッシュ。"""
    global _gpu_static
    if _gpu_static is None:
        out = _run_ps(PS_GPU_STATIC)
        try:
            d = json.loads(out)
            adps = d.get("adapters") or []
            if isinstance(adps, dict):
                adps = [adps]
            _gpu_static = {"ram": int(d.get("ram", 0) or 0), "adapters": adps}
        except Exception:
            _gpu_static = {"ram": 0, "adapters": []}
    return _gpu_static


def get_adapter_usage() -> list:
    """[{luid, dedicated, shared}, ...]＝GPUごとの専用/共有 使用バイト（Task Managerの「メモリ」相当）。"""
    return get_vram()[1]


def compute_igpu(adapters: list, nvidia_used_bytes, static: dict):
    """内蔵GPU(GPU1)の全体メモリ使用状況を返す。無ければNone。
    ・nvidia-smiの使用量に最も近い『専用使用』のluid＝NVIDIA(5070)とみなして除外
    ・残りの実アダプタ(=使用量最大)を内蔵GPUとし、専用+共有 を合算（仮想ディスプレイ等は使用ほぼ0で自然に外れる）
    ・分母＝専用総容量(レジストリ) + 共有上限(RAMの半分)＝Task Managerと同じ枠組み。"""
    # 実在判定は『レジストリ基準』＝NVIDIA以外で専用VRAM総量を持つアダプタ＝内蔵GPU。
    # ⇒ モニターをグラボ側に挿して内蔵が遊んでても（使用ほぼ0でも）欄を消さない（2026-06-17修正）。
    others = [x for x in static.get("adapters", [])
              if "NVIDIA" not in (x.get("desc", "") or "").upper() and x.get("dedtotal", 0)]
    if not others:
        return None                       # 内蔵GPUが無い機種（単一dGPU）は欄を出さない
    pick = min(others, key=lambda x: x.get("dedtotal", 1 << 62))  # 専用が小さい＝内蔵
    ded_total = pick.get("dedtotal", 0)
    name = pick.get("desc", "内蔵GPU")
    shared_total = (static.get("ram", 0) or 0) // 2
    if ded_total + shared_total <= 0:
        return None
    # 使用量：NVIDIA(=nvidia-smi使用量に最も近いluid)を除外し、残りの非NVIDIAで最大を内蔵に充てる（遊んでたら0）。
    ded_used = shared_used = 0
    real = [a for a in (adapters or []) if (a.get("dedicated", 0) + a.get("shared", 0)) > 0]
    nvidia_luid = None
    if nvidia_used_bytes and real:
        nvidia_luid = min(real, key=lambda a: abs(a.get("dedicated", 0) - nvidia_used_bytes)).get("luid")
    non_nv = [a for a in real if a.get("luid") != nvidia_luid]
    if non_nv:
        ig = max(non_nv, key=lambda a: a.get("dedicated", 0) + a.get("shared", 0))
        ded_used, shared_used = ig.get("dedicated", 0), ig.get("shared", 0)
    return {"name": name, "used": ded_used + shared_used, "total": ded_total + shared_total,
            "ded_used": ded_used, "shared_used": shared_used,
            "ded_total": ded_total, "shared_total": shared_total}


def nvidia_shared_used(totals, adapters) -> int:
    """NVIDIAアダプタが共有メモリ(システムRAM)を借りてる量(bytes)。VRAM溢れの検知用。
    luidはnvidia-smiの使用量に最も近い『専用使用』で同定する。"""
    if not totals or not adapters:
        return 0
    nv_used = totals.get("used", 0) * 1024**2
    reals = [a for a in adapters if (a.get("dedicated", 0) + a.get("shared", 0)) > 0]
    if not reals:
        return 0
    nv = min(reals, key=lambda a: abs(a.get("dedicated", 0) - nv_used))
    return nv.get("shared", 0)


# ─────────────────────────────────────────────────────────
#  Ollama：ロード中モデルの確認と、優しい解放
# ─────────────────────────────────────────────────────────
def ollama_ps() -> list | None:
    """ロード中モデル一覧。Noneは未接続。[]はロードなし。"""
    try:
        with urllib.request.urlopen(f"{OLLAMA_URL}/api/ps", timeout=3) as r:
            return json.loads(r.read().decode("utf-8")).get("models", [])
    except Exception:
        return None


def ollama_unload(model: str) -> bool:
    """keep_alive=0 を送ってモデルだけVRAMから降ろす（サーバーは生かしたまま）。"""
    body = json.dumps({"model": model, "keep_alive": 0}).encode("utf-8")
    req = urllib.request.Request(f"{OLLAMA_URL}/api/generate", data=body,
                                 headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            r.read()
        return True
    except Exception:
        return False


# ─────────────────────────────────────────────────────────
#  画像生成(SD)：Forge / ComfyUI のVRAMを優しく解放
#  ・Forge  : POST /sdapi/v1/unload-checkpoint（--api起動が必要）
#  ・ComfyUI: POST /free {"unload_models":true,"free_memory":true}
#  ※未起動/未導入なら status は None ＝ 繋がらない扱い（エラーにしない）
# ─────────────────────────────────────────────────────────
def forge_status() -> dict | None:
    """Forge接続確認＋ロード中チェックポイント名。Noneは未接続。"""
    try:
        with urllib.request.urlopen(f"{FORGE_URL}/sdapi/v1/options", timeout=3) as r:
            d = json.loads(r.read().decode("utf-8"))
        return {"ckpt": d.get("sd_model_checkpoint") or t("model_unknown")}
    except Exception:
        return None


def forge_unload() -> bool:
    req = urllib.request.Request(f"{FORGE_URL}/sdapi/v1/unload-checkpoint",
                                 data=b"", method="POST")
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            r.read()
        return True
    except Exception:
        return False


def comfy_status() -> dict | None:
    """ComfyUI接続確認（/system_stats）。Noneは未接続/未導入。"""
    try:
        with urllib.request.urlopen(f"{COMFY_URL}/system_stats", timeout=3) as r:
            d = json.loads(r.read().decode("utf-8"))
        devs = d.get("devices", []) or []
        free = sum(x.get("vram_free", 0) for x in devs)
        return {"vram_free_gb": (free / 1024**3) if free else None}
    except Exception:
        return None


def comfy_free() -> bool:
    body = json.dumps({"unload_models": True, "free_memory": True}).encode("utf-8")
    req = urllib.request.Request(f"{COMFY_URL}/free", data=body,
                                 headers={"Content-Type": "application/json"}, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            r.read()
        return True
    except Exception:
        return False


# ─────────────────────────────────────────────────────────
#  プロセス終了（手放すの“力技”版）
# ─────────────────────────────────────────────────────────
def kill_pid(pid: int) -> bool:
    try:
        r = subprocess.run(["taskkill", "/PID", str(pid), "/F"],
                           capture_output=True, text=True, creationflags=NOWIN)
        return r.returncode == 0
    except Exception:
        return False


# ─────────────────────────────────────────────────────────
#  アプリごとの使用GPU指定（Windowsの「グラフィック設定」と同じ）
#    0=自動(Windows任せ) / 1=省電力GPU(=内蔵AMD) / 2=高パフォーマンス(=RTX5070)
#    HKCU\Software\Microsoft\DirectX\UserGpuPreferences に exeパスごとに保存。
#    ※効くのは描画系VRAMだけ。CUDA(LLM/SD)やdwm等は対象外。反映はアプリ再起動後。
# ─────────────────────────────────────────────────────────
GPU_PREF_KEY = r"Software\Microsoft\DirectX\UserGpuPreferences"


def pref_label(p):
    """GPU指定の表示名（言語連動）。0=自動 / 1=内蔵 / 2=5070。"""
    return {0: t("pref_label_auto"), 1: t("pref_label_igpu"),
            2: t("pref_label_5070")}.get(p, t("pref_label_auto"))


def get_gpu_pref(exe_path: str | None) -> int:
    if not exe_path:
        return 0
    try:
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, GPU_PREF_KEY) as k:
            val, _ = winreg.QueryValueEx(k, exe_path)
        m = re.search(r"GpuPreference=(\d+)", val or "")
        return int(m.group(1)) if m else 0
    except OSError:
        return 0


def set_gpu_pref(exe_path: str | None, pref: int) -> bool:
    if not exe_path:
        return False
    try:
        k = winreg.CreateKey(winreg.HKEY_CURRENT_USER, GPU_PREF_KEY)
        try:
            if pref == 0:
                try:
                    winreg.DeleteValue(k, exe_path)   # 自動＝エントリ削除
                except FileNotFoundError:
                    pass
            else:
                winreg.SetValueEx(k, exe_path, 0, winreg.REG_SZ, f"GpuPreference={pref};")
        finally:
            winreg.CloseKey(k)
        return True
    except OSError:
        return False


# ─────────────────────────────────────────────────────────
#  判定ユーティリティ
# ─────────────────────────────────────────────────────────
def classify(name: str | None) -> str:
    """'ai' / 'maybe' / 'other' / 'unknown'"""
    if not name:
        return "unknown"
    n = name.lower()
    if any(k in n for k in AI_KEYWORDS):
        return "ai"
    if n in PY_NAMES:
        return "maybe"
    return "other"


def is_protected(name: str | None, pid: int) -> bool:
    if pid is not None and pid < 10:
        return True
    if not name:
        return True  # 名前が取れない＝システム系の可能性大。安全側に倒す
    return name.lower() in PROTECTED


def gb(mib_or_bytes, is_bytes=False) -> float:
    return mib_or_bytes / 1024**3 if is_bytes else mib_or_bytes / 1024


def usage_color(pct: float) -> str:
    if pct >= 95:
        return C_CRIT
    if pct >= 90:
        return C_HIGH
    if pct >= 80:
        return C_WARN
    return C_OK


def short_gpu_name(name: str | None) -> str:
    """'NVIDIA GeForce RTX 5070' → '5070' のように短い呼び名にする。"""
    if not name:
        return "GPU"
    m = re.search(r"(?:RTX|GTX|RX|Arc)\s*([A-Za-z]*\d{3,4}\w*)", name, re.I)
    if m:
        return m.group(1)
    return name.split()[-1]


# ─────────────────────────────────────────────────────────
#  記録(スナップショット)まわりの補助
# ─────────────────────────────────────────────────────────


CHART_TOP_N = 10            # 推移グラフに描くプロセス数の上限（ピークVRAMの多い順）
AUTO_FLUSH_EVERY = 5        # 自動記録：この回数ごと＋停止時にだけ実ファイルへ書き出す（I/O削減）


def fmt_start(s: str | None) -> str:
    """PowerShellのStartTime(ISO8601)を 'YYYY-MM-DD HH:MM:SS' に整形。失敗時は素のまま。"""
    if not s:
        return ""
    m = re.match(r"(\d{4}-\d{2}-\d{2})T(\d{2}:\d{2}:\d{2})", s)
    return f"{m.group(1)} {m.group(2)}" if m else s


def build_gtag(adapters, labels, rank) -> str:
    """プロセスが『どのGPUに何GB乗ってるか』の短いタグを作る（一覧表示と記録で共通）。"""
    adps = [a for a in (adapters or []) if a.get("luid") in labels]
    adps.sort(key=lambda a: rank.get(a["luid"], 99))   # 順位固定（表示が揺れない）
    if len(adps) <= 1:
        # 1GPUだけ: ラベルのみ（量は消費容量列と同じなので省略／0個なら空文字）
        return "".join(("🍃" if labels[a["luid"]] == t("pref_label_igpu") else "🖥") + labels[a["luid"]]
                       for a in adps)
    # 複数GPUに乗ってる(dwm等): どっちに何GB乗ってるか数字で＝肩代わり判定用。
    # 表示は実数(lbytes＝実際に乗ってる分)。luid識別用の bytes(Dedicated)とは別。
    return " ・".join(
        f'{"🍃" if labels[a["luid"]] == t("pref_label_igpu") else "🖥"}{labels[a["luid"]]} '
        f'{a.get("lbytes", a.get("bytes", 0)) / 1024**3:.2f}G'
        for a in adps)


def get_pid_monitor_map() -> dict:
    """{pid: 'モニター名'} を返す。可視ウィンドウを持つプロセスだけ拾える。
    ・複数ウィンドウがあれば一番大きいウィンドウのモニターを採用
    ・追加ライブラリ不要(ctypesのみ)。Windows以外/失敗時は空dict
    ※Ollama等の窓なしプロセスは入らない＝呼び出し側で「ウィンドウなし」にする。"""
    if os.name != "nt":
        return {}
    try:
        import ctypes
        from ctypes import wintypes
    except Exception:
        return {}

    user32 = ctypes.windll.user32

    class MONITORINFOEXW(ctypes.Structure):
        _fields_ = [("cbSize", wintypes.DWORD),
                    ("rcMonitor", wintypes.RECT),
                    ("rcWork", wintypes.RECT),
                    ("dwFlags", wintypes.DWORD),
                    ("szDevice", wintypes.WCHAR * 32)]

    WNDENUMPROC = ctypes.WINFUNCTYPE(wintypes.BOOL, wintypes.HWND, wintypes.LPARAM)
    user32.EnumWindows.argtypes = [WNDENUMPROC, wintypes.LPARAM]
    user32.IsWindowVisible.argtypes = [wintypes.HWND]
    user32.GetWindowRect.argtypes = [wintypes.HWND, ctypes.POINTER(wintypes.RECT)]
    user32.GetWindowThreadProcessId.argtypes = [wintypes.HWND, ctypes.POINTER(wintypes.DWORD)]
    user32.MonitorFromWindow.restype = wintypes.HANDLE
    user32.MonitorFromWindow.argtypes = [wintypes.HWND, wintypes.DWORD]
    user32.GetMonitorInfoW.argtypes = [wintypes.HANDLE, ctypes.POINTER(MONITORINFOEXW)]

    MONITOR_DEFAULTTONEAREST = 2
    MONITORINFOF_PRIMARY = 1

    def label_of(hwnd):
        h = user32.MonitorFromWindow(hwnd, MONITOR_DEFAULTTONEAREST)
        mi = MONITORINFOEXW()
        mi.cbSize = ctypes.sizeof(MONITORINFOEXW)
        if not user32.GetMonitorInfoW(h, ctypes.byref(mi)):
            return t("monitor_q")
        m = re.search(r"DISPLAY(\d+)", mi.szDevice or "")
        name = t("monitor_n", n=m.group(1)) if m else (mi.szDevice or t("monitor_q"))
        if mi.dwFlags & MONITORINFOF_PRIMARY:
            name += t("monitor_primary")
        return name

    best = {}  # pid -> (ウィンドウ面積, モニター名)

    def cb(hwnd, lparam):
        if not user32.IsWindowVisible(hwnd):
            return True
        rect = wintypes.RECT()
        if not user32.GetWindowRect(hwnd, ctypes.byref(rect)):
            return True
        area = (rect.right - rect.left) * (rect.bottom - rect.top)
        if area <= 0:
            return True
        pid = wintypes.DWORD()
        user32.GetWindowThreadProcessId(hwnd, ctypes.byref(pid))
        p = pid.value
        if p not in best or area > best[p][0]:
            best[p] = (area, label_of(hwnd))
        return True

    try:
        user32.EnumWindows(WNDENUMPROC(cb), 0)
    except Exception:
        return {}
    return {p: lbl for p, (a, lbl) in best.items()}


# ─────────────────────────────────────────────────────────
#  GUI
# ─────────────────────────────────────────────────────────
class VRAMaPyon(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title(t("app_title"))
        self.geometry("660x980")
        self.minsize(580, 600)
        self.configure(fg_color=BG)
        ctk.set_appearance_mode("light")

        self._busy = False
        self.auto = ctk.BooleanVar(value=True)
        self.ai_only = ctk.BooleanVar(value=False)
        self.scale_ratio = ctk.BooleanVar(value=True)   # True=📊全体比率(GPU比) / False=📈細かい上下(自分比)

        # プロセス一覧は「使い回し」する（毎回破棄しないのでちらつき・スクロール戻りなし）
        self._rows = {}            # pid -> {frame, sig, ...各ウィジェット}
        self._row_order = []       # 現在パックされてるpidの並び
        self._placeholder = None   # 「取れなかった」等のメッセージ用
        self._hist = {}            # pid -> deque(VRAM GB履歴) ＝推移スパークライン用

        self.mini = None           # フロート小窓（常駐ちら見）
        self._last_g = None        # 直近のGPU情報（小窓を開いた瞬間に即描画する用）
        self._danger = False       # 小窓が危険域の明滅中か
        self._pulse_t = 0.0        # 明滅アニメの位相

        self.auto_rec = False               # 自動記録モードON/OFF
        self._auto_path = None              # 自動記録の保存先（最初に一度だけ選ぶ）
        self._auto_after = None             # 自動記録タイマーのafter id
        self._auto_busy = False             # 前回の保存がまだ走ってるか（重複書き込み防止）
        self._auto_n = 0                    # 今セッションで自動記録した回数
        self._auto_buffer = []              # 自動記録の未書き出しバッファ（数回に1度まとめて書く）
        self._rec_sec = 30                  # 自動記録の間隔（秒）＝メニュー選択の真の値

        self._build_header()
        self._build_controls()
        self._build_gauge()
        self._build_ollama_panel()
        self._build_sd_panel()
        self._build_proc_list()

        self.after(200, self.refresh)

    # ---- ヘッダー ----
    def _build_header(self):
        head = ctk.CTkFrame(self, fg_color=PINK, corner_radius=0, height=70)
        head.pack(fill="x")
        head.pack_propagate(False)
        ctk.CTkLabel(head, text=t("header_title"), font=("Yu Gothic UI", 24, "bold"),
                     text_color="white").pack(side="left", padx=20)
        ctk.CTkLabel(head, text=t("header_sub"),
                     font=SM_FONT, text_color=PINK_SOFT).pack(side="right", padx=20)

    # ---- 操作バー ----
    def _build_controls(self):
        bar = ctk.CTkFrame(self, fg_color=BG)
        bar.pack(fill="x", padx=16, pady=(8, 4))
        ctk.CTkButton(bar, text=t("btn_refresh"), width=80, font=APP_BOLD,
                      fg_color=PINK, hover_color=PINK_DARK, text_color="white",
                      corner_radius=16, command=self.refresh).pack(side="left")
        ctk.CTkSwitch(bar, text=t("sw_auto"), variable=self.auto, font=APP_FONT,
                      progress_color=PINK, command=self._on_auto).pack(side="left", padx=14)
        ctk.CTkSwitch(bar, text=t("sw_ai_only"), variable=self.ai_only, font=APP_FONT,
                      progress_color=LAVENDER, command=self.refresh).pack(side="left")
        # 日英トグル（ゴースト調＝塗りのアクションボタンと競合させない）
        ctk.CTkButton(bar, text=t("lang_btn"), width=80, font=SM_FONT,
                      fg_color="transparent", hover_color="#F3D7E6", text_color=PINK_DARK,
                      border_width=1, border_color="#EEC9DD", corner_radius=16,
                      command=self.toggle_language).pack(side="left", padx=(14, 0))
        # 隅に常駐してちら見できる小窓へ
        ctk.CTkButton(bar, text=t("btn_mini"), width=72, font=APP_BOLD,
                      fg_color=LAVENDER, hover_color="#7468C8", text_color="white",
                      corner_radius=16, command=self._open_mini).pack(side="right")

        # 記録の行＝混み合い防止に2段目へ（手動📸／自動▶️開始・終了／間隔）
        bar2 = ctk.CTkFrame(self, fg_color=BG)
        bar2.pack(fill="x", padx=16, pady=(0, 4))
        # 今の瞬間を1回だけ記録（手動スナップ）
        self._snap_btn = ctk.CTkButton(bar2, text=t("btn_snap"), width=96, font=APP_BOLD,
                                       fg_color="#E89BC4", hover_color=PINK_DARK,
                                       text_color="white", corner_radius=16,
                                       command=self._export_snapshot)
        self._snap_btn.pack(side="left")
        # 自動記録＝開始/終了トグル。ONの間、右の間隔ごとに同じファイルへ自動追記→グラフが育つ
        self._autorec_btn = ctk.CTkButton(bar2, text=t("autorec_start"), width=156, font=APP_BOLD,
                                          fg_color=LAVENDER, hover_color="#7468C8",
                                          text_color="white", corner_radius=16,
                                          command=self._toggle_autorec)
        self._autorec_btn.pack(side="left", padx=(8, 0))
        # 間隔は「数字＋単位」を1つのメニューに（小窓化してもつぶれず読める）
        u = t("rec_unit")
        self._rec_menu = ctk.CTkOptionMenu(bar2, values=[f"15{u}", f"30{u}", f"60{u}"],
                                           width=82, font=SM_FONT, fg_color="#E7DEF3",
                                           button_color=LAVENDER, button_hover_color="#7468C8",
                                           text_color=INK, command=self._on_interval)
        self._rec_menu.set(f"{self._rec_sec}{u}")
        self._rec_menu.pack(side="left", padx=(8, 0))
        self._sync_autorec_btn()

    # ---- GPU全体ゲージ ----
    def _build_gauge(self):
        card = ctk.CTkFrame(self, fg_color=CARD, corner_radius=18)
        card.pack(fill="x", padx=16, pady=6)
        self.gpu_name = ctk.CTkLabel(card, text=t("gauge_loading"), font=SM_FONT,
                                     text_color=SUB)
        self.gpu_name.pack(anchor="w", padx=18, pady=(12, 0))
        self.gauge_label = ctk.CTkLabel(card, text=t("gauge_empty"), font=BIG_FONT,
                                        text_color=INK)
        self.gauge_label.pack(anchor="w", padx=18)
        self.bar = ctk.CTkProgressBar(card, height=16, corner_radius=8,
                                      progress_color=C_OK, fg_color=PINK_SOFT)
        self.bar.set(0)
        self.bar.pack(fill="x", padx=18, pady=(6, 4))
        self.gauge_sub = ctk.CTkLabel(card, text="", font=SM_FONT, text_color=SUB)
        self.gauge_sub.pack(anchor="w", padx=18, pady=(0, 8))
        # 内蔵GPU(GPU1)も 5070ゲージと同じ見た目で揃える（仕切り→名前→大きな数字→太バー→専用/共有内訳）。
        # データが取れる機種だけ _update_igpu が pack して出す（取れなければ畳む）。
        self.igpu_row = ctk.CTkFrame(card, fg_color="transparent")
        ctk.CTkFrame(self.igpu_row, fg_color="#F0E6EF", height=1).pack(fill="x", pady=(0, 8))
        self.igpu_name = ctk.CTkLabel(self.igpu_row, text=t("igpu_loading"), font=SM_FONT, text_color=SUB)
        self.igpu_name.pack(anchor="w")
        self.igpu_value = ctk.CTkLabel(self.igpu_row, text=t("gauge_empty"), font=BIG_FONT, text_color=INK)
        self.igpu_value.pack(anchor="w")
        self.igpu_bar = ctk.CTkProgressBar(self.igpu_row, height=16, corner_radius=8,
                                           progress_color=C_OK, fg_color=PINK_SOFT)
        self.igpu_bar.set(0)
        self.igpu_bar.pack(fill="x", pady=(6, 4))
        self.igpu_sub = ctk.CTkLabel(self.igpu_row, text="", font=SM_FONT, text_color=SUB)
        self.igpu_sub.pack(anchor="w")

    # ---- Ollama 解放パネル ----
    def _build_ollama_panel(self):
        self.oll_card = ctk.CTkFrame(self, fg_color="#F3ECFF", corner_radius=18)
        self.oll_card.pack(fill="x", padx=16, pady=6)
        top = ctk.CTkFrame(self.oll_card, fg_color="transparent")
        top.pack(fill="x", padx=16, pady=(10, 8))
        ctk.CTkLabel(top, text=t("oll_title"),
                     font=APP_BOLD, text_color="#5B3CA8").pack(side="left")
        # 状態をタイトル横に常時表示（未接続/モデルなしの時はこの1行だけ＝薄く畳む）
        self.oll_status = ctk.CTkLabel(top, text=t("oll_checking"), font=SM_FONT, text_color=SUB)
        self.oll_status.pack(side="left", padx=10)
        self.oll_all = ctk.CTkButton(top, text=t("oll_btn_all"), width=80, font=SM_FONT,
                                     fg_color=LAVENDER, hover_color="#7468C8",
                                     text_color="white", corner_radius=14,
                                     command=self._unload_all)
        self.oll_all.pack(side="right")
        # 「速い？なぜ遅い？」はこっちの専用ツールへ（別アプリで起動）
        ctk.CTkButton(top, text=t("oll_btn_diag"), width=96, font=SM_FONT,
                      fg_color="#8E6FE0", hover_color="#7A5BD0", text_color="white",
                      corner_radius=14, command=self._open_diagnosis).pack(side="right", padx=(0, 8))
        # モデル行の本体。ロード中モデルがある時だけ pack して展開する
        self.oll_body = ctk.CTkFrame(self.oll_card, fg_color="transparent")

    # ---- SD(Forge/ComfyUI) 解放パネル ----
    def _build_sd_panel(self):
        self.sd_card = ctk.CTkFrame(self, fg_color="#FCEAF4", corner_radius=18)
        self.sd_card.pack(fill="x", padx=16, pady=6)
        self.sd_title = ctk.CTkLabel(self.sd_card, text=t("sd_title"),
                                     font=APP_BOLD, text_color="#A23A77", anchor="w")
        self.sd_title.pack(anchor="w", padx=16, pady=(10, 8))
        # Forge/ComfyUI行の本体。繋がってる時だけ pack（未接続なら薄く畳む）
        self.sd_body = ctk.CTkFrame(self.sd_card, fg_color="transparent")

    # ---- プロセス一覧 ----
    def _build_proc_list(self):
        head = ctk.CTkFrame(self, fg_color="transparent")
        head.pack(fill="x", padx=22, pady=(6, 0))
        ctk.CTkLabel(head, text=t("proc_title"),
                     font=APP_BOLD, text_color=INK).pack(side="left")
        ctk.CTkLabel(head, text=t("proc_note"),
                     font=SM_FONT, text_color=SUB).pack(side="left", padx=(10, 0))
        # 足型グラフの縦軸モード切り替え（普段=📊全体比率／調べる時=📈細かい上下）
        self.scale_seg = ctk.CTkSegmentedButton(
            head, values=[t("seg_ratio"), t("seg_detail")], font=SM_FONT,
            command=self._on_scale_mode, height=28, corner_radius=12,
            fg_color="#F6DCEB", selected_color=PINK, selected_hover_color=PINK_DARK,
            unselected_color="#F6DCEB", unselected_hover_color="#EEC9DD", text_color=INK)
        self.scale_seg.set(t("seg_ratio"))
        self.scale_seg.pack(side="right")
        self.proc_frame = ctk.CTkScrollableFrame(self, fg_color=CARD, corner_radius=18)
        self.proc_frame.pack(fill="both", expand=True, padx=16, pady=(6, 16))

    # ─────────────────────────────────────────────
    #  記録（スナップショット）
    #   1行＝1プロセスのフラット表。スナップショット情報(取得時刻/GPU/合計VRAM)は
    #   各行の先頭列に展開＝Excelで並べ替え/フィルタ/ピボットがそのまま効く。
    #   同じファイルを選べば「追記」＝時系列でためて後から比較できる。本命は.xlsx。
    # ─────────────────────────────────────────────
    SNAP_WIDTHS = [20, 22, 14, 14, 14, 10, 10, 8, 26, 9, 8, 14, 20, 12, 20, 52]

    def _export_snapshot(self):
        from tkinter import filedialog
        import datetime
        stamp = datetime.datetime.now().strftime("%Y%m%d")         # 既定名は日付＝同じ日は1ファイルに溜まる
        path = filedialog.asksaveasfilename(
            title=t("snap_dlg_title"),
            defaultextension=".xlsx", initialfile=t("snap_default_name", stamp=stamp),
            confirmoverwrite=False,                # 追記運用なので上書き確認は出さない
            filetypes=[(t("snap_ft_excel"), "*.xlsx"), (t("snap_ft_csv"), "*.csv")])
        if not path:
            return
        self._snap_btn.configure(state="disabled", text=t("snap_saving"))
        threading.Thread(target=self._do_export, args=(path,), daemon=True).start()

    def _collect_rows(self):
        """現在のスナップショットを「1行＝1プロセス」のデータ行に。手動/自動の両記録で共通。"""
        import datetime
        totals = get_gpu_totals()
        procs = get_proc_vram()
        mon = get_pid_monitor_map()
        labels, rank = self._resolve_gpu_labels(procs, totals)
        ai_only = self.ai_only.get()           # 画面と同じフィルタで「表示したもの」を残す
        now = datetime.datetime.now()
        flt = t("flt_ai") if ai_only else t("flt_all")
        gname = totals["name"] if totals else t("totals_na")
        g_tot = gb(totals["total"]) if totals else None
        g_use = gb(totals["used"]) if totals else None
        g_free = gb(totals["free"]) if totals else None
        g_pct = (totals["used"] / totals["total"] * 100) if (totals and totals["total"]) else None
        g_util = totals["util"] if totals else None

        rows = []              # 各行＝1プロセス（先頭8列にスナップショット情報を毎行展開）
        for d in procs:
            kind = classify(d.get("name"))
            if ai_only and kind not in ("ai", "maybe"):
                continue
            rows.append([
                now, gname, g_tot, g_use, g_free, g_pct, g_util, flt,
                d.get("name") or t("proc_unknown"), t(f"kind_{kind}"), d.get("pid"),
                round(gb(d.get("bytes", 0), is_bytes=True), 3),
                build_gtag(d.get("adapters"), labels, rank),
                mon.get(d.get("pid"), t("win_none")),
                fmt_start(d.get("start")), d.get("path") or "",
            ])
        return rows

    def _save_rows(self, path, rows):
        """rowsを拡張子に応じてxlsx/csvへ追記。通算データ行数を返す。"""
        if path.lower().endswith(".csv"):
            return self._append_csv(path, rows)
        return self._append_xlsx(path, rows)

    def _do_export(self, path):
        try:
            rows = self._collect_rows()
            total = self._save_rows(path, rows)
            self.after(0, lambda: self._export_done(True, "", path, len(rows), total))
        except Exception as e:
            msg = str(e)
            self.after(0, lambda: self._export_done(False, msg, path, 0, 0))

    def _append_xlsx(self, path, rows):
        """既存があれば追記、無ければ見出し＋書式付きで新規作成。通算データ行数を返す。"""
        try:
            import openpyxl
        except ImportError:                        # customtkinterと同じく必要時に自動導入
            subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"],
                           check=True, creationflags=NOWIN)
            import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
            ws = wb.worksheets[0]              # データシートは必ず1枚目（グラフシートを掴まない）
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = t("snap_sheet")
            ws.append(t("snap_headers"))
            for c in range(1, len(t("snap_headers")) + 1):
                cell = ws.cell(row=1, column=c)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="EA4C9D")
                cell.alignment = Alignment(vertical="center")
            ws.freeze_panes = "A2"                 # 見出し固定
            for i, w in enumerate(self.SNAP_WIDTHS, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

        for rv in rows:
            ws.append(rv)
            rr = ws.max_row
            ws.cell(rr, 1).number_format = "yyyy-mm-dd hh:mm:ss"
            for col in (3, 4, 5, 12):
                ws.cell(rr, col).number_format = "0.00"
            for col in (6, 7):
                ws.cell(rr, col).number_format = "0.0"
            ws.cell(rr, 11).number_format = "0"

        last = get_column_letter(len(t("snap_headers")))
        ws.auto_filter.ref = f"A1:{last}{ws.max_row}"   # オートフィルタを全範囲に
        self._add_databar(ws)               # プロセスVRAM列をセル内バーで見やすく
        self._rebuild_pivot(wb, ws)         # 別シートに時系列の折れ線グラフ3種を作り直す
        wb.save(path)
        return ws.max_row - 1

    @staticmethod
    def _add_databar(ws):
        """プロセスVRAM列(L)をセル内データバー(ピンク)に。毎回クリアして貼り直す。"""
        try:
            from openpyxl.formatting.rule import DataBarRule
            from openpyxl.formatting.formatting import ConditionalFormattingList
            ws.conditional_formatting = ConditionalFormattingList()
            if ws.max_row >= 2:
                rule = DataBarRule(start_type="num", start_value=0,
                                   end_type="max", color="EA4C9D")
                ws.conditional_formatting.add(f"L2:L{ws.max_row}", rule)
        except Exception:
            pass

    @staticmethod
    def _rebuild_pivot(wb, data_ws=None):
        """データシートを時系列ピボット化し、別シートに折れ線グラフ3種を作り直す。
        ・GPU全体VRAM(使用/合計) ・GPU使用率% ・プロセス別VRAM(ピーク上位)。
        言語が変わってもデータシート以外を全消し→再生成するので確実。"""
        from openpyxl.chart import LineChart, Reference
        data = data_ws or wb.worksheets[0]
        for s in list(wb.worksheets):           # データシート以外(=過去のグラフシート)を撤去
            if s is not data:
                del wb[s.title]

        # 列インデックス固定: 1=時刻 3=合計 4=使用 6=使用率 9=プロセス名 12=プロセスVRAM
        snaps, order = {}, []
        for r in range(2, data.max_row + 1):
            ts = data.cell(r, 1).value
            if ts is None:
                continue
            s = snaps.get(ts)
            if s is None:
                s = {"used": data.cell(r, 4).value, "total": data.cell(r, 3).value,
                     "pct": data.cell(r, 6).value, "procs": {}}
                snaps[ts] = s
                order.append(ts)
            nm = data.cell(r, 9).value or "?"
            s["procs"][nm] = s["procs"].get(nm, 0) + (data.cell(r, 12).value or 0)
        if not order:
            return
        order.sort()

        peak = {}                               # プロセスはピークVRAMの多い順で上位だけ（見やすさ優先）
        for s in snaps.values():
            for nm, v in s["procs"].items():
                peak[nm] = max(peak.get(nm, 0), v)
        top = [nm for nm, _ in sorted(peak.items(), key=lambda kv: kv[1], reverse=True)[:CHART_TOP_N]]

        ws = wb.create_sheet(t("chart_sheet"))
        # グラフを上(目立つ位置)に、ピボット表はその下(DATA0行〜)に置く＝開いてすぐ折れ線が見える
        DATA0 = 50
        head = [t("chart_col_time"), t("chart_col_used"),
                t("chart_col_total"), t("chart_col_pct")] + top
        for c, val in enumerate(head, start=1):
            ws.cell(DATA0, c, val)
        labels, rr = [], DATA0
        for ts in order:
            rr += 1
            s = snaps[ts]
            # X軸は「文字列ラベル」にする＝Excelが等間隔カテゴリ軸で扱い、点が潰れない
            # (日時/数値参照だと日付・数値軸になり、近接時刻の点が1か所に固まって線が見えない)
            label = ts.strftime("%m-%d %H:%M:%S") if hasattr(ts, "strftime") else str(ts)
            labels.append(label)
            vals = [label, s["used"], s["total"], s["pct"]] + [round(s["procs"].get(nm, 0), 3) for nm in top]
            for c, val in enumerate(vals, start=1):
                ws.cell(rr, c, val)
        last = rr
        ws.column_dimensions["A"].width = 16

        # カテゴリ(X軸)は文字列参照(strRef)＋キャッシュを手動で当てる。
        # ※set_categories()はnumRef(数値参照)を作るため、文字列セルでも数値扱い→全点が左に潰れる。
        from openpyxl.chart.data_source import AxDataSource, StrRef, StrData, StrVal
        catf = f"'{ws.title}'!$A${DATA0 + 1}:$A${last}"

        def add_line(title, c0, c1, y_title, anchor, style):
            ch = LineChart()
            ch.title = title
            ch.style = style
            ch.height, ch.width = 7.5, 20
            ch.y_axis.title = y_title
            ref = Reference(ws, min_col=c0, max_col=c1, min_row=DATA0, max_row=last)
            ch.add_data(ref, titles_from_data=True)   # DATA0行=系列名、その下=値
            cache = StrData(ptCount=len(labels), pt=[StrVal(idx=i, v=v) for i, v in enumerate(labels)])
            for s in ch.series:                       # 各系列に文字列カテゴリを明示
                s.cat = AxDataSource(strRef=StrRef(f=catf, strCache=cache))
            ws.add_chart(ch, anchor)

        add_line(t("chart_gpu_title"), 2, 3, "GB", "A1", 12)
        add_line(t("chart_pct_title"), 4, 4, "%", "A17", 13)
        if top:
            add_line(t("chart_proc_title"), 5, 4 + len(top), "GB", "A33", 10)
        wb.active = wb.index(ws)        # 開いた時に「推移グラフ」タブを前面に

    def _append_csv(self, path, rows):
        """フラットCSVに追記（無ければ見出し付き）。通算データ行数を返す。"""
        import csv, datetime
        exists = os.path.exists(path) and os.path.getsize(path) > 0
        with open(path, "a", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            if not exists:
                w.writerow(t("snap_headers"))
            for rv in rows:
                w.writerow([v.strftime("%Y-%m-%d %H:%M:%S")
                            if isinstance(v, datetime.datetime) else v for v in rv])
        with open(path, encoding="utf-8-sig") as f:
            return max(0, sum(1 for _ in f) - 1)        # 見出し1行を除く通算行数

    def _export_done(self, ok, err, path, n, total):
        self._snap_btn.configure(state="normal", text=t("btn_snap"))
        if ok:
            if messagebox.askyesno(t("app_name"), t("snap_done_open", n=n, total=total, path=path)):
                try:
                    os.startfile(path)            # 「開く」を選んだら既定アプリ(Excel)で開く
                except Exception:
                    pass
        else:
            hint = t("snap_perm_hint") if "permission" in err.lower() else ""
            messagebox.showerror(t("app_name"), t("snap_fail", err=err, hint=hint))

    # ---- 自動記録（一定間隔で同じファイルに追記＝数分でグラフが育つ）----
    def _interval_ms(self):
        return max(5, self._rec_sec) * 1000

    def _sync_autorec_btn(self):
        if not hasattr(self, "_autorec_btn"):
            return
        if self.auto_rec:
            self._autorec_btn.configure(text=t("autorec_recording", n=self._auto_n),
                                        fg_color=C_CRIT, hover_color="#C0392B")
        else:
            self._autorec_btn.configure(text=t("autorec_start"),
                                        fg_color=LAVENDER, hover_color="#7468C8")

    def _on_interval(self, _=None):
        import re
        try:                                     # メニュー表示("30秒"/"30s")から数字だけ取り出す
            self._rec_sec = max(5, int(re.sub(r"\D", "", self._rec_menu.get()) or "30"))
        except Exception:
            self._rec_sec = 30
        if self.auto_rec and self._auto_after:   # 記録中なら次回を新間隔で組み直す
            self.after_cancel(self._auto_after)
            self._auto_after = self.after(self._interval_ms(), self._auto_tick)

    def _toggle_autorec(self):
        if self.auto_rec:
            self._stop_autorec()
            return
        from tkinter import filedialog
        import datetime
        stamp = datetime.datetime.now().strftime("%Y%m%d")
        path = filedialog.asksaveasfilename(
            title=t("autorec_dlg_title"),
            defaultextension=".xlsx", initialfile=t("snap_default_name", stamp=stamp),
            confirmoverwrite=False,
            filetypes=[(t("snap_ft_excel"), "*.xlsx"), (t("snap_ft_csv"), "*.csv")])
        if not path:
            return
        self._auto_path = path
        self._auto_n = 0
        self._auto_buffer = []
        self.auto_rec = True
        self._sync_autorec_btn()
        self._auto_tick()              # 即1点目を記録

    def _stop_autorec(self):
        self.auto_rec = False
        if self._auto_after:
            self.after_cancel(self._auto_after)
            self._auto_after = None
        self._sync_autorec_btn()
        if self._auto_buffer:                # 停止時に残りバッファをまとめて書き出す
            threading.Thread(target=self._stop_flush, daemon=True).start()

    def _stop_flush(self):
        try:
            self._flush_auto()
        except Exception as e:
            self.after(0, lambda: messagebox.showwarning(t("app_name"), t("autorec_err", err=str(e))))

    def _auto_tick(self):
        if not self.auto_rec:
            return
        if not self._auto_busy:        # 前回の処理が長引いてたら今回はスキップ（多重実行防止）
            self._auto_busy = True
            threading.Thread(target=self._auto_save, daemon=True).start()
        self._auto_after = self.after(self._interval_ms(), self._auto_tick)

    def _auto_save(self):
        # 毎回はデータをバッファに貯めるだけ＝軽い。実ファイル書き込み＆グラフ再構築は
        # 数回に1度(AUTO_FLUSH_EVERY)＋停止時だけ＝長時間回してもファイルI/Oで重くならない。
        try:
            self._auto_buffer.extend(self._collect_rows())
            self._auto_n += 1
            if self._auto_n == 1 or self._auto_n % AUTO_FLUSH_EVERY == 0:
                self._flush_auto()
            self.after(0, self._sync_autorec_btn)
        except Exception as e:
            msg = str(e)
            self.after(0, lambda: self._auto_done_err(msg))
        finally:
            self._auto_busy = False

    def _flush_auto(self):
        """バッファのぶんをまとめてファイルへ追記＋グラフ再構築。失敗時はバッファを戻す。"""
        if not self._auto_buffer:
            return
        buf = self._auto_buffer
        self._auto_buffer = []
        try:
            self._save_rows(self._auto_path, buf)
        except Exception:
            self._auto_buffer = buf + self._auto_buffer   # 失敗したら戻して次回再挑戦
            raise

    def _auto_done_err(self, msg):
        self._stop_autorec()
        messagebox.showwarning(t("app_name"), t("autorec_err", err=msg))

    # ─────────────────────────────────────────────
    #  更新まわり
    # ─────────────────────────────────────────────
    def _on_auto(self):
        if self.auto.get():
            self.refresh()

    def _on_scale_mode(self, value=None):
        # 📊全体比率(GPU比) ↔ 📈細かい上下(自分比) を切り替えて全行を即描き直す（データ取り直し不要）
        self.scale_ratio.set(value == t("seg_ratio"))
        for row in self._rows.values():
            self._draw_cell(row["cell"])

    def toggle_language(self):
        # 🌐 日英を切り替えて全UIを作り直す（履歴 self._hist は残してスパークラインを維持）
        global LANG
        LANG = "en" if LANG == "ja" else "ja"
        self._close_mini()
        for w in self.winfo_children():
            w.destroy()
        self._rows = {}
        self._row_order = []
        self._placeholder = None
        self._busy = False
        self.title(t("app_title"))
        self._build_header()
        self._build_controls()
        self._build_gauge()
        self._build_ollama_panel()
        self._build_sd_panel()
        self._build_proc_list()
        self.refresh()

    def refresh(self):
        if self._busy:
            return
        self._busy = True
        threading.Thread(target=self._collect, daemon=True).start()

    def _collect(self):
        totals = get_gpu_totals()
        procs, adapters = get_vram()       # 1回のPS呼び出しでプロセス別＋アダプタ別をまとめて取得
        oll = ollama_ps()
        forge = forge_status()
        comfy = comfy_status()
        self.after(0, lambda: self._apply(totals, procs, adapters, oll, forge, comfy))

    def _apply(self, totals, procs, adapters, oll, forge, comfy):
        static = get_gpu_static()
        self._update_gauge(totals, adapters, static)
        self._update_igpu(adapters, totals)
        self._update_ollama(oll)
        self._update_sd(forge, comfy)
        self._update_procs(procs, totals)
        self._busy = False
        if self.auto.get():
            self.after(REFRESH_MS, self.refresh)

    def _update_gauge(self, g, adapters=None, static=None):
        self._last_g = g
        if not g:
            self.gpu_name.configure(text=t("gauge_na"))
            self.gauge_label.configure(text=t("gauge_empty"))
            self.bar.set(0)
            self.gauge_sub.configure(text="")
            self._render_mini(g)
            return
        pct = g["used"] / g["total"] * 100 if g["total"] else 0
        col = usage_color(pct)
        self.gpu_name.configure(text=f"🎮 {g['name']}")
        self.gauge_label.configure(
            text=f"{gb(g['used']):.1f} / {gb(g['total']):.1f} GB  ({pct:.0f}%)",
            text_color=col)
        self.bar.configure(progress_color=col)
        self.bar.set(min(pct / 100, 1.0))
        mark = {C_OK: t("mark_ok"), C_WARN: t("mark_warn"),
                C_HIGH: t("mark_high"), C_CRIT: t("mark_crit")}[col]
        # 内蔵ゲージと表示を揃える：専用(=カードのVRAM) / 共有(=RAMから借りる枠)。
        # 5070の共有が増える＝VRAMが溢れてシステムRAMに逃げた合図（要注意）。
        ded_used, ded_total = gb(g["used"]), gb(g["total"])
        shared_total = ((static.get("ram", 0) if static else 0) or 0) / 2 / 1024**3
        shared_used = nvidia_shared_used(g, adapters or []) / 1024**3
        self.gauge_sub.configure(
            text=f"{mark}   " + t("sub_dedshared", du=ded_used, dt=ded_total,
                                  su=shared_used, st=shared_total))
        self._render_mini(g)

    @staticmethod
    def _short_igpu(name):
        # 5070の正式名表示（NVIDIA GeForce RTX 5070）に揃えて、内蔵もベンダー名＋（内蔵）で出す
        n = (name or t("igpu_fallback")).replace("(TM)", "").replace("(R)", "")
        n = " ".join(n.split())
        return f"{n}{t('igpu_suffix')}"

    def _update_igpu(self, adapters, totals):
        """内蔵GPU(GPU1)のゲージを更新（5070ゲージと同じ見た目＝使用率カラー＋専用/共有の内訳）。
        全体の％は専用+共有の合算（=Task Managerの『GPU メモリ』）。専用だけ満タンでも合算で余裕なら緑。"""
        nvidia_used = (totals["used"] * 1024**2) if totals else None
        info = compute_igpu(adapters or [], nvidia_used, get_gpu_static())
        if not info or info["total"] <= 0:
            self.igpu_row.pack_forget()
            return
        used_gb = info["used"] / 1024**3
        total_gb = info["total"] / 1024**3
        pct = (info["used"] / info["total"] * 100) if info["total"] else 0
        col = usage_color(pct)
        mark = {C_OK: t("mark_ok"), C_WARN: t("mark_warn"),
                C_HIGH: t("mark_high"), C_CRIT: t("mark_crit")}[col]
        self.igpu_name.configure(text=f"🍃 {self._short_igpu(info['name'])}")
        self.igpu_value.configure(text=f"{used_gb:.1f} / {total_gb:.1f} GB  ({pct:.0f}%)",
                                  text_color=col)
        self.igpu_bar.configure(progress_color=col)
        self.igpu_bar.set(min(pct / 100, 1.0))
        self.igpu_sub.configure(
            text=f"{mark}   " + t("sub_dedshared",
                du=info['ded_used']/1024**3, dt=info['ded_total']/1024**3,
                su=info['shared_used']/1024**3, st=info['shared_total']/1024**3))
        if not self.igpu_row.winfo_ismapped():
            self.igpu_row.pack(fill="x", padx=18, pady=(0, 10))

    # ---- フロート小窓（常駐ちら見）----
    def _open_mini(self):
        if self.mini:                      # すでに小窓化してたら何もしない
            return
        self.withdraw()                    # 本体は引っ込める
        m = ctk.CTkToplevel(self)
        m.title(t("app_name"))
        m.overrideredirect(True)           # 枠なしのちっちゃい窓に
        m.attributes("-topmost", True)     # 常に最前面
        m.configure(fg_color=PINK)
        w, h = 212, 70
        m.geometry(f"{w}x{h}+{m.winfo_screenwidth() - w - 24}+48")  # 右上に出す

        inner = ctk.CTkFrame(m, fg_color=CARD, corner_radius=14)
        inner.pack(fill="both", expand=True, padx=3, pady=3)
        toprow = ctk.CTkFrame(inner, fg_color="transparent")
        toprow.pack(fill="x", padx=12, pady=(8, 2))
        self.mini_top = ctk.CTkLabel(toprow, text="🐾 -- / -- GB", font=APP_BOLD, text_color=INK)
        self.mini_top.pack(side="left")
        self.mini_pct = ctk.CTkLabel(toprow, text="--%", font=APP_BOLD, text_color=INK)
        self.mini_pct.pack(side="right")
        self.mini_bar = ctk.CTkProgressBar(inner, height=10, corner_radius=6,
                                           progress_color=C_OK, fg_color=PINK_SOFT)
        self.mini_bar.set(0)
        self.mini_bar.pack(fill="x", padx=12, pady=(0, 8))

        self.mini = m
        self._danger = False
        for wdg in (m, inner, toprow, self.mini_top, self.mini_pct):
            wdg.bind("<Button-1>", self._mini_press)
            wdg.bind("<B1-Motion>", self._mini_move)
            wdg.bind("<Double-Button-1>", lambda e: self._restore_from_mini())
            wdg.bind("<Button-3>", self._mini_menu)
        self._render_mini(self._last_g)

    def _render_mini(self, g):
        if not self.mini:
            return
        if not g:
            self.mini_top.configure(text=t("mini_na"))
            self.mini_pct.configure(text="--%")
            self.mini_bar.set(0)
            self._danger = False
            self.mini.configure(fg_color=PINK)
            return
        pct = g["used"] / g["total"] * 100 if g["total"] else 0
        col = usage_color(pct)
        self.mini_top.configure(text=f"🐾 {gb(g['used']):.1f} / {gb(g['total']):.1f} GB")
        self.mini_pct.configure(text=f"{pct:.0f}%", text_color=col)
        self.mini_bar.configure(progress_color=col)
        self.mini_bar.set(min(pct / 100, 1.0))
        # 危険域(90%~)は枠をふわっと赤く明滅。80%~はアンバー、普段はピンク。
        if pct >= 90:
            if not self._danger:
                self._danger = True
                self._pulse_t = 0.0
                self._pulse()
        else:
            self._danger = False
            self.mini.configure(fg_color=(C_WARN if pct >= 80 else PINK))

    @staticmethod
    def _lerp_hex(a, b, t):
        a, b = a.lstrip("#"), b.lstrip("#")
        c = [round(int(a[i:i + 2], 16) + (int(b[i:i + 2], 16) - int(a[i:i + 2], 16)) * t)
             for i in (0, 2, 4)]
        return "#{:02x}{:02x}{:02x}".format(*c)

    def _pulse(self):
        # 危険域の間だけ枠を赤↔淡赤でゆらゆら（ふわっと呼吸）
        if not self.mini or not self._danger:
            return
        self._pulse_t = (self._pulse_t + 0.05) % 2.0
        t = self._pulse_t if self._pulse_t <= 1 else 2 - self._pulse_t
        self.mini.configure(fg_color=self._lerp_hex(C_CRIT, "#F7C0BB", t))
        self.after(70, self._pulse)

    def _mini_press(self, e):
        self._mini_off = (e.x_root - self.mini.winfo_x(), e.y_root - self.mini.winfo_y())

    def _mini_move(self, e):
        self.mini.geometry(f"+{e.x_root - self._mini_off[0]}+{e.y_root - self._mini_off[1]}")

    def _mini_menu(self, e):
        menu = Menu(self.mini, tearoff=0)
        menu.add_command(label=t("mini_open"), command=self._restore_from_mini)
        menu.add_command(label=t("mini_quit"), command=self.destroy)
        try:
            menu.tk_popup(e.x_root, e.y_root)
        finally:
            menu.grab_release()

    def _restore_from_mini(self):
        self._close_mini()
        self.deiconify()
        self.lift()

    def _close_mini(self):
        if self.mini:
            self.mini.destroy()
            self.mini = None

    def _update_ollama(self, oll):
        for w in self.oll_body.winfo_children():
            w.destroy()
        if oll is None:
            self.oll_status.configure(text=t("oll_disconnected"), text_color=SUB)
            self.oll_all.configure(state="disabled")
            self.oll_body.pack_forget()                  # 薄く畳む
            return
        if not oll:
            self.oll_status.configure(text=t("oll_none"), text_color=SUB)
            self.oll_all.configure(state="disabled")
            self.oll_body.pack_forget()                  # 薄く畳む
            return
        self.oll_status.configure(text=t("oll_loaded", n=len(oll)), text_color="#5B3CA8")
        self.oll_all.configure(state="normal")
        self.oll_body.pack(fill="x", padx=16, pady=(0, 10))   # ある時だけ展開
        for m in oll:
            name = m.get("name") or m.get("model", "?")
            sv = m.get("size_vram", 0)
            st = m.get("size", 0) or sv          # 総サイズ。無ければVRAM量で代用
            vram_gb = sv / 1024**3
            ratio = (sv / st * 100) if st > 0 else 100
            if ratio >= 90:
                fit_txt, fit_col = t("fit_allgpu"), C_OK
            elif ratio >= 50:
                fit_txt, fit_col = t("fit_ramover", p=100 - ratio), C_HIGH
            else:
                fit_txt, fit_col = t("fit_mostram"), C_CRIT
            row = ctk.CTkFrame(self.oll_body, fg_color="transparent")
            row.pack(fill="x", pady=2)
            ctk.CTkLabel(row, text=f"💭 {name}", font=APP_FONT,
                         text_color="#4A2F86", anchor="w").pack(side="left")
            ctk.CTkLabel(row, text=f"{vram_gb:.1f} GB", font=APP_BOLD,
                         text_color="#5B3CA8").pack(side="left", padx=8)
            ctk.CTkLabel(row, text=fit_txt, font=SM_FONT,
                         text_color=fit_col).pack(side="left", padx=4)
            ctk.CTkButton(row, text=t("btn_release"), width=64, font=SM_FONT,
                          fg_color=LAVENDER, hover_color="#7468C8", text_color="white",
                          corner_radius=12,
                          command=lambda n=name: self._unload_one(n)).pack(side="right")

    def _unload_one(self, name):
        ok = ollama_unload(name)
        if not ok:
            messagebox.showwarning(t("app_name"), t("msg_unload_fail", name=name))
        self.refresh()

    def _unload_all(self):
        oll = ollama_ps() or []
        for m in oll:
            ollama_unload(m.get("name") or m.get("model", ""))
        self.refresh()

    def _open_diagnosis(self):
        """別アプリ「Ollama診断GUI」を起動（速い？なぜ遅い？はそっちで）。"""
        path = find_diagnosis_tool()
        if not path:
            # 候補のうち分かりやすい場所(兄弟フォルダ想定)を案内に出す
            here = os.path.dirname(os.path.abspath(__file__))
            hint = os.path.normpath(os.path.join(here, "..", DIAG_DIRNAMES[0], DIAG_FILENAME))
            messagebox.showwarning(t("app_name"), t("msg_diag_notfound", path=hint))
            return
        try:
            subprocess.Popen([sys.executable, path], creationflags=NOWIN)
        except Exception as e:
            messagebox.showwarning(t("app_name"), t("msg_diag_launchfail", e=e))

    # ---- SD(Forge/ComfyUI) 解放まわり ----
    def _update_sd(self, forge, comfy):
        for w in self.sd_body.winfo_children():
            w.destroy()
        connected = []
        if forge is not None:
            connected.append(("Forge", t("sd_ckpt", name=forge['ckpt']), self._free_forge))
        if comfy is not None:
            sub = (t("sd_vramfree", gb=comfy['vram_free_gb'])
                   if comfy.get("vram_free_gb") else t("sd_connected"))
            connected.append(("ComfyUI", sub, self._free_comfy))
        if not connected:                       # 未接続＝1行に畳む（普段はこれで省スペース）
            self.sd_title.configure(text=t("sd_disconnected"), text_color=SUB)
            self.sd_body.pack_forget()
            return
        self.sd_title.configure(text=t("sd_title"), text_color="#A23A77")
        self.sd_body.pack(fill="x", padx=16, pady=(0, 10))
        for name, sub, cmd in connected:
            self._sd_row(name, True, sub, cmd)

    def _sd_row(self, name, on, sub, cmd):
        row = ctk.CTkFrame(self.sd_body, fg_color="transparent")
        row.pack(fill="x", pady=2)
        left = ctk.CTkFrame(row, fg_color="transparent")
        left.pack(side="left", fill="x", expand=True)
        ctk.CTkLabel(left, text=f"{'🎨' if on else '・'} {name}", font=APP_FONT,
                     text_color=("#8A2E66" if on else SUB), anchor="w").pack(anchor="w")
        ctk.CTkLabel(left, text=sub, font=SM_FONT, text_color=SUB, anchor="w").pack(anchor="w")
        btn = ctk.CTkButton(row, text=t("btn_release"), width=64, font=SM_FONT,
                            fg_color="#E48ABF", hover_color="#D46FAE", text_color="white",
                            corner_radius=12, command=cmd)
        if not on:
            btn.configure(state="disabled")
        btn.pack(side="right")

    def _free_forge(self):
        if not messagebox.askyesno(t("app_name"), t("msg_forge_confirm")):
            return
        if not forge_unload():
            messagebox.showwarning(t("app_name"), t("msg_forge_fail"))
        self.refresh()

    def _free_comfy(self):
        if not comfy_free():
            messagebox.showwarning(t("app_name"), t("msg_comfy_fail"))
        self.refresh()

    def _resolve_gpu_labels(self, procs, totals):
        """luid → (表示名, 並び順位)。NVIDIA(5070)のluidを特定し、ほかは内蔵扱い。

        ⚠️NVIDIA判定は **実数(Local)合計が nvidia-smi 使用量に最も近いluid** で行う。
        旧版は「Dedicated(予約)最大のluid＝NVIDIA」だったが、NVIDIA Overlay 等が内蔵側luidに
        巨大な予約(数十GB)を積むと内蔵をNVIDIAと誤認しラベルが逆転した（Forge=実数5GBが5070
        なのに「内蔵」表示になる等）。実数はnvidia-smiの物理使用と一致するので、これで騙されない。
        順位は実数合計の多い順で固定＝タグを安定表示。"""
        loc = {}    # luid -> Local(実数) 合計
        ded = {}    # luid -> Dedicated(予約) 合計（フォールバック用）
        for d in procs:
            for a in (d.get("adapters") or []):
                lu = a.get("luid")
                loc[lu] = loc.get(lu, 0) + a.get("lbytes", 0)
                ded[lu] = ded.get(lu, 0) + a.get("bytes", 0)
        for tbl in (loc, ded):
            tbl.pop(None, None)
        luids = set(loc) | set(ded)
        if not luids:
            return {}, {}
        # NVIDIAのluid＝実数合計が nvidia-smi 使用量に最も近いもの（幽霊予約に騙されない）
        nv = None
        if totals and totals.get("used"):
            target = totals["used"] * 1024 * 1024            # MB → bytes
            nv = min(luids, key=lambda lu: abs(loc.get(lu, 0) - target))
        if nv is None:                                       # nvidia-smi無し時は実数最大をメイン扱い
            nv = max(luids, key=lambda lu: loc.get(lu, 0))
        rest = sorted((lu for lu in luids if lu != nv),
                      key=lambda lu: loc.get(lu, 0), reverse=True)
        order = [nv] + rest
        labels = {order[0]: (short_gpu_name(totals["name"]) if totals else "GPU")}
        for i, lu in enumerate(order[1:]):
            labels[lu] = t("pref_label_igpu") if i == 0 else f"GPU{i + 1}"
        rank = {lu: i for i, lu in enumerate(order)}
        return labels, rank

    @staticmethod
    def _pid_text(pid, gtag):
        return t("pid_text", pid=pid, gtag=gtag) if gtag else t("pid_text_plain", pid=pid)

    # ---- プロセス一覧の更新（行は破棄せず使い回す＝ちらつき・スクロール戻りなし）----
    def _update_procs(self, procs, totals):
        labels, rank = self._resolve_gpu_labels(procs, totals)

        # フィルタ＆並び（VRAM降順のまま）。各 item に表示用の確定値を詰める
        desired = []
        for d in procs:
            name = d.get("name")
            pid = d.get("pid")
            kind = classify(name)
            if self.ai_only.get() and kind not in ("ai", "maybe"):
                continue
            gtag = build_gtag(d.get("adapters"), labels, rank)
            desired.append(dict(pid=pid, name=name, by=d.get("bytes", 0),
                                ded=d.get("dedBytes", 0),
                                kind=kind, path=d.get("path"), gtag=gtag))

        # 空・エラー時はメッセージだけ
        if not procs:
            self._clear_rows()
            self._show_placeholder(t("ph_noproc"))
            return
        if not desired:
            self._clear_rows()
            self._show_placeholder(t("ph_noai"))
            return
        self._remove_placeholder()

        pos = self._get_scroll()                  # スクロール位置を退避
        wanted = [it["pid"] for it in desired]
        wanted_set = set(wanted)
        changed = False

        # 消えたプロセスの行を片付け
        for pid in list(self._rows):
            if pid not in wanted_set:
                self._rows[pid]["frame"].destroy()
                del self._rows[pid]
                self._hist.pop(pid, None)         # 消えたプロセスの履歴も捨てる
                changed = True

        # 追加・更新（構成が同じなら中身だけ書き換え＝破棄しない）
        for it in desired:
            pid = it["pid"]
            sig = (it["kind"], is_protected(it["name"], pid), bool(it["path"]))
            row = self._rows.get(pid)
            if row and row["sig"] == sig:
                self._update_row(row, it)
            else:
                if row:
                    row["frame"].destroy()
                self._rows[pid] = self._build_row(it, sig)
                changed = True

        # 増減・並び替えが起きた時だけ詰め直す（毎回はしない）
        did_repack = False
        if changed or wanted != self._row_order:
            for pid in wanted:
                self._rows[pid]["frame"].pack_forget()
            for pid in wanted:
                self._rows[pid]["frame"].pack(fill="x", padx=6, pady=3)
            self._row_order = wanted
            did_repack = True

        if pos and did_repack:                    # 詰め直した時だけ位置を戻す
            self.after_idle(lambda: self._set_scroll(pos))

    def _build_row(self, it, sig):
        pid = it["pid"]
        name = it["name"]
        kind = it["kind"]
        path = it["path"]
        prot = sig[1]
        badge = {"ai": "🤖", "maybe": "🐍", "other": "・", "unknown": "🔒"}[kind]
        bgc = PINK_SOFT if kind == "ai" else ("#F6F0FF" if kind == "maybe" else "#F7F7F9")
        disp = name or t("proc_unknown")

        frame = ctk.CTkFrame(self.proc_frame, fg_color=bgc, corner_radius=12)
        # 3列: [左セル=名前/PID＋背景グラフ(伸縮)] [GB(固定)] [操作(固定)]
        # 左セルはweight=1で全行同じ幅＝グラフの長さがそろう。グラフは文字の後ろにうっすら。
        frame.grid_columnconfigure(0, weight=1, minsize=180)
        frame.grid_columnconfigure(1, minsize=90)
        frame.grid_columnconfigure(2, minsize=206)

        # 左セル：1枚のCanvasに「背景グラフ」→「名前・PID文字」を重ね描き（_draw_cell）
        cell = Canvas(frame, height=CELL_H, bg=bgc, highlightthickness=0, bd=0)
        cell.grid(row=0, column=0, sticky="nsew", padx=(10, 0), pady=3)
        cell.bind("<Configure>", lambda e, c=cell: self._draw_cell(c))

        # GB列（右寄せ・固定幅）— ここが全行で揃う
        gb_lbl = ctk.CTkLabel(frame, text="", width=82, anchor="e",
                              font=("Yu Gothic UI", 15, "bold"),
                              text_color=PINK_DARK if kind == "ai" else INK)
        gb_lbl.grid(row=0, column=1, sticky="e", padx=(0, 6))

        # 操作列（右から: 終了/🔒 → GPUボタン → 指定タグ）固定幅なのでGB位置がブレない
        actions = ctk.CTkFrame(frame, fg_color="transparent")
        actions.grid(row=0, column=2, sticky="e", padx=(0, 10))
        pref_lbl = None
        if prot:
            ctk.CTkLabel(actions, text=t("lbl_protected"), font=SM_FONT, text_color=SUB).pack(
                side="right", padx=6)
        else:
            ctk.CTkButton(actions, text=t("btn_kill"), width=52, font=SM_FONT,
                          fg_color=C_CRIT, hover_color="#C8372F", text_color="white",
                          corner_radius=12,
                          command=lambda: self._kill(name, pid)).pack(side="right", padx=(4, 0))
            if path:
                ctk.CTkButton(actions, text=t("btn_gpu"), width=58, font=SM_FONT,
                              fg_color=LAVENDER, hover_color="#7468C8", text_color="white",
                              corner_radius=12,
                              command=lambda: self._gpu_dialog(disp, path)).pack(side="right", padx=2)
                pref_lbl = ctk.CTkLabel(actions, text="", width=68, anchor="e",
                                        font=SM_FONT, text_color=PINK)
                pref_lbl.pack(side="right", padx=2)

        row = dict(frame=frame, sig=sig, cell=cell, gb_lbl=gb_lbl,
                   pref_lbl=pref_lbl, path=path, badge=badge, name=disp, kind=kind)
        self._update_row(row, it)
        return row

    def _update_row(self, row, it):
        # VRAM推移を1サンプル足す
        h = self._hist.get(it["pid"])
        if h is None:
            h = self._hist[it["pid"]] = deque(maxlen=SPARK_N)
        h.append(it["by"] / 1024**3)
        # 足型の状態（張り付き/増加/安定）を判定 → GB数字と肉球・線の色に反映
        total_gb = gb(self._last_g["total"]) if (self._last_g and self._last_g.get("total")) else None
        state = self._proc_state(h, total_gb)
        if state == "pinned":
            gb_col = C_CRIT
        elif row["kind"] == "ai":
            gb_col = PINK_DARK
        else:
            gb_col = INK
        row["gb_lbl"].configure(text=f"{it['by'] / 1024**3:.2f} GB", text_color=gb_col)
        # 左セルに描く内容を更新して再描画（背景グラフ＋名前/PID文字）
        cell = row["cell"]
        cell._series = h
        cell._state = state
        cell._name = f'{row["badge"]} {row["name"]}'
        sub = self._pid_text(it["pid"], it["gtag"])
        # 予約(Dedicated)が実数(Local)よりハッキリ多い時だけ「・予約○G」を併記（普段は出さない）。
        # NVIDIA Overlay 等が「実 0.05／予 31.9」のように一目で分かるように。
        if (it.get("ded", 0) or 0) - (it.get("by", 0) or 0) > RESV_GAP_BYTES:
            sub += t("resv_tag", gb=(it["ded"]) / 1024**3)
        cell._sub = sub
        self._draw_cell(cell)
        if row["pref_lbl"] is not None:
            # ボタンを押さなくても 自動/固定 が分かるよう常に表示
            pref = get_gpu_pref(row["path"])
            if pref == 0:
                row["pref_lbl"].configure(text=t("pref_auto"), text_color=SUB)
            else:
                row["pref_lbl"].configure(text=t("pref_fixed", label=pref_label(pref)),
                                          text_color=(LAVENDER if pref == 1 else PINK))

    def _proc_state(self, series, total_gb):
        """足型の状態を返す: 'pinned'(天井張り付き) / 'rising'(持続増加) / 'stable'。
        瞬間のスパイクで誤爆しないよう、前半平均と後半平均の比較で“持続”だけ拾う。"""
        s = list(series) if series else []
        if not s:
            return "stable"
        last = s[-1]
        if total_gb and (last / total_gb) >= PIN_SHARE:        # GPUを大きく占有していて…
            recent = s[-min(len(s), 8):]
            avg = sum(recent) / len(recent)
            if (max(recent) - min(recent)) <= max(avg * PIN_FLAT, 0.05):  # …高止まり＝張り付き
                return "pinned"
        if len(s) >= 6:
            half = len(s) // 2
            early, late = s[:half], s[half:]
            em = sum(early) / len(early)
            lm = sum(late) / len(late)
            if lm > em * RISE_RATIO and last >= em:            # 後半が持続的に上＆まだ下がってない
                return "rising"
        return "stable"

    @staticmethod
    def _halo_text(cv, x, y, text, font, fill, anchor="w"):
        """白フチ付きテキスト（背景グラフの上でも文字が読めるように8方向の白＋本体）。"""
        for dx in (-1, 0, 1):
            for dy in (-1, 0, 1):
                if dx or dy:
                    cv.create_text(x + dx, y + dy, text=text, font=font,
                                   fill="white", anchor=anchor)
        cv.create_text(x, y, text=text, font=font, fill=fill, anchor=anchor)

    def _draw_cell(self, cv):
        """左セルを描画：背景にVRAM推移グラフ（あしあとライン＋🐾）、前面に名前・PID文字（白フチ）。
        ・縦軸2モード: 📊全体比率=0〜GPU総量（張り付くと線が“上に貼り付く”）／📈自分比=履歴の最小〜最大（細かい上下）
        ・状態色: 🔥張り付き=赤 / ▲増加=オレンジ / 安定=ピンク（線・肉球＋右下タグ）"""
        cv.delete("all")
        w = cv.winfo_width()
        h = CELL_H
        series = getattr(cv, "_series", None)
        state = getattr(cv, "_state", "stable")

        g = self._last_g
        total_gb = gb(g["total"]) if (g and g.get("total")) else None
        ratio_mode = self.scale_ratio.get() and (total_gb is not None)

        if state == "pinned":
            c_line, c_fill, c_paw = SPARK_RED_LINE, SPARK_RED_FILL, C_CRIT
        elif state == "rising":
            c_line, c_fill, c_paw = SPARK_BG_LINE, SPARK_BG_FILL, C_HIGH
        else:
            c_line, c_fill, c_paw = SPARK_BG_LINE, SPARK_BG_FILL, PINK

        if w > 1 and series:
            s = list(series)
            pr, pt, pb = 8, 7, 6
            maxpts = max(2, int((w - pr) / SPARK_STEP) + 1)
            s = s[-maxpts:]
            n = len(s)
            if ratio_mode:                         # 0〜GPU総量で固定スケール＝占有率がそのまま高さに
                lo, hi = 0.0, total_gb
                cv.create_line(2, pt - 2, w - 2, pt - 2, fill=SPARK_CEIL, dash=(2, 3))  # 天井(満タン)
            else:                                  # そのプロセス自身の最小〜最大＝細かい上下が見える
                lo, hi = min(s), max(s)
            rng = (hi - lo) or 1.0

            def X(i):                              # 最新(i=n-1)を右端に、古いほど左へ等間隔
                return (w - pr) - (n - 1 - i) * SPARK_STEP

            def Y(v):
                yy = h - pb - ((v - lo) / rng) * (h - pt - pb)
                return max(pt - 2, min(h - pb, yy))   # 天井張り付きでハミ出さないようクランプ

            pts = [(X(i), Y(v)) for i, v in enumerate(s)]
            if n > 1:
                base = h - pb
                poly = [(pts[0][0], base)] + pts + [(pts[-1][0], base)]
                cv.create_polygon([c for p in poly for c in p], fill=c_fill, outline="")
                cv.create_line([c for p in pts for c in p], fill=c_line, width=1.5,
                               smooth=True, capstyle="round", joinstyle="round")
            lx, ly = pts[-1]                                  # 先端＝最新サンプル
            cv.create_oval(lx - 3.2, ly - 1.4, lx + 3.2, ly + 3.4, fill=c_paw, outline="")
            for dx, dy in ((-3.1, -3.3), (0, -4.3), (3.1, -3.3)):
                cv.create_oval(lx + dx - 1.2, ly + dy - 1.2, lx + dx + 1.2, ly + dy + 1.2,
                               fill=c_paw, outline="")

        # 前面：名前(上)・PID/GPU内訳(下)を白フチ付きで重ね描き（波の上でも読める）
        self._halo_text(cv, 2, 14, getattr(cv, "_name", ""), APP_BOLD, INK)
        self._halo_text(cv, 2, 32, getattr(cv, "_sub", ""), SM_FONT, SUB)
        # 状態タグ（右下）: 🔥張り付き / ▲増加
        tag = {"pinned": t("tag_pinned"), "rising": t("tag_rising"), "stable": ""}.get(state, "")
        if tag and w > 4:
            self._halo_text(cv, w - 4, 33, tag, ("Yu Gothic UI", 10, "bold"),
                            C_CRIT if state == "pinned" else C_HIGH, anchor="e")

    def _clear_rows(self):
        for row in self._rows.values():
            row["frame"].destroy()
        self._rows.clear()
        self._row_order = []

    def _show_placeholder(self, text):
        if self._placeholder is None:
            self._placeholder = ctk.CTkLabel(self.proc_frame, font=SM_FONT,
                                             text_color=SUB, justify="left")
            self._placeholder.pack(anchor="w", padx=12, pady=12)
        self._placeholder.configure(text=text)

    def _remove_placeholder(self):
        if self._placeholder is not None:
            self._placeholder.destroy()
            self._placeholder = None

    def _get_scroll(self):
        try:
            return self.proc_frame._parent_canvas.yview()
        except Exception:
            return None

    def _set_scroll(self, pos):
        try:
            self.proc_frame._parent_canvas.yview_moveto(pos[0])
        except Exception:
            pass

    def _kill(self, name, pid):
        if not messagebox.askyesno(
                t("app_name"),
                t("msg_kill_confirm", name=name, pid=pid)):
            return
        if not kill_pid(pid):
            messagebox.showwarning(t("app_name"), t("msg_kill_fail"))
        self.refresh()

    # ---- アプリごとの使用GPU指定 ----
    def _gpu_dialog(self, name, path):
        cur = get_gpu_pref(path)
        win = ctk.CTkToplevel(self)
        win.title(t("gpu_dlg_title"))
        # 本体の少し内側に開く（画面外や裏に出ないように位置を固定）
        win.geometry(f"380x384+{self.winfo_rootx() + 80}+{self.winfo_rooty() + 110}")
        win.configure(fg_color=BG)
        win.transient(self)
        # CTkToplevelはWindowsで本体の裏に開きがち→確実に前面へ出す
        win.attributes("-topmost", True)
        win.after(10, win.lift)
        win.after(30, win.focus_force)
        win.after(400, lambda: win.attributes("-topmost", False))  # 常時最前面は解除
        win.after(150, lambda: self._safe_grab(win))               # 前面化の後にモーダル化

        ctk.CTkLabel(win, text=t("gpu_dlg_head", name=name), font=APP_BOLD, text_color=INK,
                     wraplength=340).pack(pady=(16, 0), padx=16)
        ctk.CTkLabel(win, text=t("gpu_dlg_current", label=pref_label(cur)), font=APP_BOLD,
                     text_color=PINK_DARK).pack(pady=(0, 6))

        opts = [
            (2, t("gpu_opt_5070"), PINK, PINK_DARK),
            (1, t("gpu_opt_igpu"), LAVENDER, "#7468C8"),
            (0, t("gpu_opt_auto"), "#B8B0BC", "#9A9098"),
        ]
        for pref, label, col, hov in opts:
            is_cur = (pref == cur)   # いま選ばれてる項目は ✓＋太枠＋太字で目立たせる
            ctk.CTkButton(win, text=(f"✓ {label}" if is_cur else label),
                          font=(APP_BOLD if is_cur else APP_FONT), height=42,
                          fg_color=col, hover_color=hov, text_color="white",
                          corner_radius=14,
                          border_width=(3 if is_cur else 0), border_color=INK,
                          command=lambda p=pref: self._apply_pref(win, name, path, p)
                          ).pack(fill="x", padx=20, pady=5)

        ctk.CTkLabel(win, text=t("gpu_dlg_note"),
                     font=("Yu Gothic UI", 10), text_color=SUB,
                     justify="left").pack(pady=(8, 12), padx=18)

    @staticmethod
    def _safe_grab(win):
        try:
            win.grab_set()
        except Exception:
            pass

    def _apply_pref(self, win, name, path, pref):
        ok = set_gpu_pref(path, pref)
        try:
            win.destroy()
        except Exception:
            pass
        if ok:
            messagebox.showinfo(t("app_name"),
                                t("msg_pref_set", name=name, label=pref_label(pref)))
        else:
            messagebox.showwarning(t("app_name"), t("msg_pref_fail"))
        self.refresh()


# ─────────────────────────────────────────────────────────
#  確認用：GUIなしでデータだけ表示
# ─────────────────────────────────────────────────────────
def probe():
    print("== GPU全体 ==")
    g = get_gpu_totals()
    if g:
        print(f"  {g['name']}  {gb(g['used']):.1f}/{gb(g['total']):.1f} GB "
              f"({g['used']/g['total']*100:.0f}%)  util {g['util']}%")
    else:
        print("  取得不可")

    print("\n== プロセス別VRAM ==")
    procs = get_proc_vram()
    # luid→GPU名（VRAM最大のluidをメインGPUに）
    tot = {}
    for d in procs:
        for a in d.get("adapters", []):
            tot[a.get("luid")] = tot.get(a.get("luid"), 0) + a.get("bytes", 0)
    tot.pop(None, None)
    order = sorted(tot, key=tot.get, reverse=True)
    labels = {}
    if order:
        labels[order[0]] = short_gpu_name(g["name"]) if g else "GPU"
        for i, lu in enumerate(order[1:]):
            labels[lu] = t("pref_label_igpu") if i == 0 else f"GPU{i + 1}"
    for d in procs[:15]:
        k = classify(d.get("name"))
        tag = {"ai": "[AI]", "maybe": "[py]", "other": "    ", "unknown": "[?] "}[k]
        prot = " (保護)" if is_protected(d.get("name"), d.get("pid")) else ""
        adps = d.get("adapters") or []
        prim = max(adps, key=lambda a: a.get("bytes", 0)).get("luid") if adps else None
        gpu = labels.get(prim, "?")
        print(f"  {tag} {(d.get('name') or '?'):<20} PID {d.get('pid'):<7} "
              f"{d['bytes']/1024**3:6.2f} GB  [{gpu}]{prot}")

    print("\n== Ollama ロード中モデル ==")
    oll = ollama_ps()
    if oll is None:
        print("  未接続")
    elif not oll:
        print("  ロードなし")
    else:
        for m in oll:
            print(f"  {m.get('name')}  {m.get('size_vram',0)/1024**3:.1f} GB")


if __name__ == "__main__":
    if "--probe" in sys.argv:
        probe()
    else:
        VRAMaPyon().mainloop()
